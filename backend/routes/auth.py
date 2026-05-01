from fastapi import APIRouter, Depends, Request, Response, UploadFile, File
from datetime import datetime, timezone, timedelta
import uuid
import os
import httpx
from passlib.context import CryptContext

from deps import get_tenant_db
from helpers import get_current_user, get_school_settings_doc
from utils.audit import log_audit

router = APIRouter()
_pwd = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Cookie security — set COOKIE_SECURE=false in .env for plain HTTP local servers
_COOKIE_SECURE = os.environ.get('COOKIE_SECURE', 'true').lower() != 'false'
_COOKIE_SAMESITE = 'none' if _COOKIE_SECURE else 'lax'


async def _get_settings(db):
    """Return the canonical school settings doc, preferring the one with onboarding_complete=True."""
    s = await db.school_settings.find_one({"onboarding_complete": True}, {"_id": 0})
    return s or await db.school_settings.find_one({}, {"_id": 0})

def _set_session_cookie(response, token: str, cross_subdomain: bool = False):
    """Set session cookie. If cross_subdomain=True and in production, sets Domain for subdomain sharing."""
    cookie_kwargs = dict(
        key='session_token', value=token,
        httponly=True, secure=_COOKIE_SECURE,
        samesite=_COOKIE_SAMESITE, path='/',
        max_age=7 * 24 * 3600,
    )
    # In production, set Domain=.welltrack.com.au so cookie works across subdomains
    if cross_subdomain and os.environ.get('APP_ENV', 'production') == 'production':
        base_domain = os.environ.get('BASE_DOMAIN', 'welltrack.com.au')
        cookie_kwargs['domain'] = f'.{base_domain}'
    response.set_cookie(**cookie_kwargs)


@router.post("/auth/login-email")
async def login_email(data: dict, response: Response, db=Depends(get_tenant_db)):
    from fastapi import HTTPException
    email = data.get("email", "").lower().strip()
    password = data.get("password", "")

    settings = await _get_settings(db)
    # Default email_auth to True — if the field was never explicitly saved, allow login
    if not settings or settings.get("email_auth_enabled", True) is False:
        raise HTTPException(status_code=403, detail="Email login is not enabled. Go to Settings → Integrations to enable it.")

    user_doc = await db.users.find_one({"email": email}, {"_id": 0})
    if not user_doc or not user_doc.get("password_hash"):
        raise HTTPException(status_code=401, detail="Invalid email or password")

    if not _pwd.verify(password, user_doc["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")

    session_token = f"sess_{uuid.uuid4().hex}"
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.user_sessions.insert_one({
        "user_id": user_doc["user_id"], "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": now_iso,
    })
    await db.users.update_one({"user_id": user_doc["user_id"]}, {"$set": {"last_login": now_iso}})
    _set_session_cookie(response, session_token)
    # Determine where to send the user — always go to dashboard if setup is done
    is_complete = bool(settings and settings.get("onboarding_complete")) or \
                  await db.users.count_documents({}) > 1
    return {"message": "Login successful", "redirect": "dashboard" if is_complete else "onboarding"}


@router.put("/auth/change-password")
async def change_password(data: dict, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    from fastapi import HTTPException
    current_pw = data.get("current_password", "")
    new_pw = data.get("new_password", "")
    if len(new_pw) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")
    user_doc = await db.users.find_one({"user_id": user["user_id"]}, {"_id": 0})
    existing_hash = user_doc.get("password_hash") if user_doc else None
    if existing_hash and not _pwd.verify(current_pw, existing_hash):
        raise HTTPException(status_code=401, detail="Current password is incorrect")
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"password_hash": _pwd.hash(new_pw)}})
    return {"message": "Password changed successfully"}


@router.post("/users/{user_id}/set-password")
async def set_user_password(user_id: str, data: dict, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    from fastapi import HTTPException
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    password = data.get("password", "")
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")
    result = await db.users.update_one({"user_id": user_id}, {"$set": {"password_hash": _pwd.hash(password)}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "Password set successfully"}


# ── Google OAuth (Multi-Tenant) ──────────────────────────────────────────────
# OAuth state is stored in control_db so the callback (which always hits the
# root domain) can resolve the tenant without needing middleware context.

@router.get("/auth/google")
async def google_login(request: Request, db=Depends(get_tenant_db)):
    from fastapi import HTTPException
    settings = await _get_settings(db)
    if settings and not settings.get("google_auth_enabled", True):
        raise HTTPException(status_code=403, detail="Google login is not enabled")

    from urllib.parse import urlencode
    from control_db import control_db

    tenant_slug = getattr(request.state, "tenant_slug", None)
    state = uuid.uuid4().hex
    await control_db.oauth_states.insert_one({
        "state": state,
        "tenant_slug": tenant_slug,
        "is_super_admin": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=10)).isoformat(),
    })
    params = urlencode({
        "response_type": "code",
        "client_id": os.environ['GOOGLE_CLIENT_ID'],
        "redirect_uri": os.environ['GOOGLE_REDIRECT_URI'],
        "scope": "openid email profile",
        "state": state,
        "prompt": "select_account",
    })
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url=f"https://accounts.google.com/o/oauth2/v2/auth?{params}")


@router.get("/auth/callback")
async def google_callback(request: Request):
    """
    Google OAuth callback — tenant-aware.
    Resolves the school from the state stored in control_db, NOT from middleware.
    This allows a single callback URL on the root domain for all schools.
    """
    from fastapi.responses import RedirectResponse
    from control_db import control_db
    from database import client

    frontend_url = os.environ['FRONTEND_URL']
    base_domain = os.environ.get("BASE_DOMAIN", "welltrack.com.au")
    app_env = os.environ.get("APP_ENV", "production")
    code = request.query_params.get("code")
    state = request.query_params.get("state")
    error = request.query_params.get("error")

    if error:
        return RedirectResponse(url=f"{frontend_url}/login?error=access_denied")
    if not code or not state:
        return RedirectResponse(url=f"{frontend_url}/login?error=auth_failed")

    # 1. Look up state in control_db (not tenant DB)
    state_doc = await control_db.oauth_states.find_one_and_delete({"state": state})
    if not state_doc:
        return RedirectResponse(url=f"{frontend_url}/login?error=auth_failed")

    expires_at_str = state_doc.get("expires_at", "")
    if expires_at_str:
        exp = datetime.fromisoformat(expires_at_str).replace(tzinfo=timezone.utc)
        if exp < datetime.now(timezone.utc):
            return RedirectResponse(url=f"{frontend_url}/login?error=auth_failed")

    # 2. Extract tenant slug
    tenant_slug = state_doc.get("tenant_slug")

    # 3. Resolve the school DB
    if not tenant_slug:
        # No tenant → reject (SA OAuth is handled separately)
        return RedirectResponse(url=f"{frontend_url}/login?error=auth_failed")

    school = await control_db.schools.find_one({"slug": tenant_slug}, {"_id": 0})
    if not school or school.get("status") in ("suspended", "archived"):
        return RedirectResponse(url=f"{frontend_url}/login?error=access_denied")

    # Check if Google auth is enabled for this school
    feature_flags = school.get("feature_flags", {})
    if feature_flags.get("google_auth") is False:
        return RedirectResponse(url=f"{frontend_url}/login?error=access_denied")

    db = client[school["db_name"]]

    # 4. Exchange code for tokens
    try:
        async with httpx.AsyncClient() as hc:
            tok_resp = await hc.post(
                "https://oauth2.googleapis.com/token",
                data={
                    "code": code,
                    "client_id": os.environ['GOOGLE_CLIENT_ID'],
                    "client_secret": os.environ['GOOGLE_CLIENT_SECRET'],
                    "redirect_uri": os.environ['GOOGLE_REDIRECT_URI'],
                    "grant_type": "authorization_code",
                },
            )
        if tok_resp.status_code != 200:
            return RedirectResponse(url=f"{frontend_url}/login?error=auth_failed")
        access_token = tok_resp.json().get("access_token")

        async with httpx.AsyncClient() as hc:
            ui_resp = await hc.get(
                "https://www.googleapis.com/oauth2/v3/userinfo",
                headers={"Authorization": f"Bearer {access_token}"},
            )
        if ui_resp.status_code != 200:
            return RedirectResponse(url=f"{frontend_url}/login?error=auth_failed")
        user_info = ui_resp.json()
    except Exception:
        return RedirectResponse(url=f"{frontend_url}/login?error=auth_failed")

    email = user_info.get('email', '').lower().strip()
    name = user_info.get('name', '')
    picture = user_info.get('picture', '')

    if not email:
        return RedirectResponse(url=f"{frontend_url}/login?error=no_email")

    # 5. Look up user in the school's DB
    existing = await db.users.find_one({"email": email}, {"_id": 0})
    user_count = await db.users.count_documents({})

    if user_count == 0:
        # First user - create with Google data as defaults
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        new_user = {"user_id": user_id, "email": email, "name": name,
                    "picture": picture, "role": "admin",
                    "created_at": datetime.now(timezone.utc).isoformat()}
        await db.users.insert_one({**new_user})
        existing = new_user
    elif not existing:
        return RedirectResponse(url=f"{frontend_url}/login?error=access_denied")
    else:
        # Existing user - preserve their WellTrack name/picture, only update if not set
        update_fields = {}
        if not existing.get("name"):
            update_fields["name"] = name
        if not existing.get("picture"):
            update_fields["picture"] = picture

        if update_fields:
            await db.users.update_one({"email": email}, {"$set": update_fields})
            existing = {**existing, **update_fields}

    # 6. Create session in school DB
    user_id = existing["user_id"]
    session_token = f"sess_{uuid.uuid4().hex}"
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.user_sessions.insert_one({
        "user_id": user_id, "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": now_iso,
    })
    await db.users.update_one({"user_id": user_id}, {"$set": {"last_login": now_iso}})

    settings_doc = await _get_settings(db)
    onboarding_complete = bool(settings_doc and settings_doc.get("onboarding_complete"))
    redirect_target = "dashboard" if onboarding_complete else "onboarding"

    # 7. Build redirect URL — in production, redirect to school subdomain
    if app_env == "production":
        redirect_url = f"https://{tenant_slug}.{base_domain}/{redirect_target}"
    else:
        redirect_url = f"{frontend_url}/{redirect_target}"

    redirect = RedirectResponse(url=redirect_url)
    _set_session_cookie(redirect, session_token, cross_subdomain=True)
    return redirect


@router.put("/auth/preferences")
async def update_preferences(data: dict, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    from fastapi import HTTPException
    allowed = {"default", "dark", "system"}
    theme = data.get("theme")
    if theme is not None and theme not in allowed:
        raise HTTPException(status_code=400, detail="Invalid theme")
    update = {}
    if theme is not None:
        update["theme"] = theme
    if update:
        await db.users.update_one({"user_id": user["user_id"]}, {"$set": update})
    return {"message": "Preferences updated"}


@router.get("/auth/me")
async def get_me(user=Depends(get_current_user)):
    return user


@router.post("/auth/logout")
async def logout(request: Request, response: Response, db=Depends(get_tenant_db)):
    token = request.cookies.get("session_token")
    if token:
        await db.user_sessions.delete_one({"session_token": token})
    response.delete_cookie("session_token", path="/", samesite=_COOKIE_SAMESITE, secure=_COOKIE_SECURE)
    return {"message": "Logged out"}


# ── Impersonation (school-side handler) ──────────────────────────────────────
@router.get("/auth/impersonate")
async def impersonate(request: Request, token: str, slug: str = ""):
    """Validate a one-time impersonation token and create an admin session.
    Works from any domain — resolves the school from control_db using the slug param."""
    from fastapi.responses import RedirectResponse
    from fastapi import HTTPException
    from control_db import control_db
    from database import client

    frontend_url = os.environ['FRONTEND_URL']
    base_domain = os.environ.get("BASE_DOMAIN", "welltrack.com.au")
    app_env = os.environ.get("APP_ENV", "production")

    if not token or not slug:
        return RedirectResponse(url=f"{frontend_url}/login?error=auth_failed")

    # Resolve school from control_db
    school = await control_db.schools.find_one({"slug": slug}, {"_id": 0})
    if not school or school.get("status") not in ("active", "trial"):
        return RedirectResponse(url=f"{frontend_url}/login?error=invalid_token")

    db = client[school["db_name"]]

    # Atomically claim the token: only an unused token matches, and the same
    # operation marks it as used. This prevents a TOCTOU race where two
    # concurrent requests with the same token could both pass the "not used"
    # check before either flips the flag.
    # Tenant scoping is implicit — `db` is this school's database, so a token
    # issued for school A cannot be redeemed against school B.
    imp_doc = await db.impersonation_tokens.find_one_and_update(
        {"token": token, "used": {"$ne": True}},
        {"$set": {"used": True, "used_at": datetime.now(timezone.utc).isoformat()}},
    )
    if not imp_doc:
        # Either no such token, already used, or wrong tenant.
        # Fetch separately just to distinguish the "expired token" UX from
        # "invalid/used", so admins know to request a fresh one.
        existing = await db.impersonation_tokens.find_one({"token": token}, {"_id": 0, "expires_at": 1, "used": 1})
        if existing and existing.get("used"):
            return RedirectResponse(url=f"{frontend_url}/login?error=token_used")
        if existing:
            exp_str = existing.get("expires_at", "")
            try:
                exp = datetime.fromisoformat(exp_str).replace(tzinfo=timezone.utc)
                if exp < datetime.now(timezone.utc):
                    return RedirectResponse(url=f"{frontend_url}/login?error=token_expired")
            except (TypeError, ValueError):
                pass
        return RedirectResponse(url=f"{frontend_url}/login?error=invalid_token")

    # Token was claimed atomically — now verify expiry. If expired, the claim
    # is wasted but at least no one else can replay it.
    expires_at_str = imp_doc.get("expires_at", "")
    if expires_at_str:
        try:
            exp = datetime.fromisoformat(expires_at_str).replace(tzinfo=timezone.utc)
        except (TypeError, ValueError):
            exp = None
        if exp and exp < datetime.now(timezone.utc):
            return RedirectResponse(url=f"{frontend_url}/login?error=token_expired")

    # Find an admin user to impersonate
    admin_user = await db.users.find_one({"role": "admin"}, {"_id": 0})
    if not admin_user:
        return RedirectResponse(url=f"{frontend_url}/login?error=no_admin")

    # Create session
    session_token = f"sess_imp_{uuid.uuid4().hex}"
    expires_at = datetime.now(timezone.utc) + timedelta(minutes=30)
    await db.user_sessions.insert_one({
        "user_id": admin_user["user_id"],
        "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "impersonated_by": imp_doc.get("super_admin_id"),
    })

    # Redirect to the school's subdomain in production
    if app_env == "production":
        redirect_url = f"https://{slug}.{base_domain}/dashboard"
    else:
        redirect_url = f"{frontend_url}/dashboard"

    redirect = RedirectResponse(url=redirect_url)
    _set_session_cookie(redirect, session_token, cross_subdomain=True)
    return redirect



@router.put("/auth/role")
async def update_role(data: dict, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    from fastapi import HTTPException
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only administrators can change user roles")
    target_user_id = data.get("user_id", user["user_id"])
    role = data.get("role")
    if role not in ["teacher", "wellbeing", "leadership", "admin", "screener", "professional"]:
        raise HTTPException(status_code=400, detail="Invalid role")
    await db.users.update_one({"user_id": target_user_id}, {"$set": {"role": role}})
    return {"message": "Role updated", "role": role}



# ── Onboarding ──────────────────────────────────────────────────────────────

@router.get("/onboarding/status")
async def get_onboarding_status(db=Depends(get_tenant_db)):
    settings = await db.school_settings.find_one({"onboarding_complete": True}, {"_id": 0})
    if not settings:
        settings = await db.school_settings.find_one({}, {"_id": 0})
    complete = bool(settings and settings.get("onboarding_complete", False))
    has_users = await db.users.count_documents({}) > 0
    return {
        "complete": complete,
        "has_users": has_users,
        "school_name": settings.get("school_name", "") if settings else "",
    }


@router.post("/onboarding/school-setup")
async def onboarding_school_setup(data: dict, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    """Multi-tenant onboarding: admin already exists (created by SA). Only saves school settings."""
    from fastapi import HTTPException
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    existing = await _get_settings(db)
    if existing and existing.get("onboarding_complete"):
        raise HTTPException(status_code=400, detail="Setup has already been completed")

    await db.school_settings.update_one({}, {"$set": {
        "school_name": data.get("school_name", "My School"),
        "school_type": data.get("school_type", "both"),
        "current_term": data.get("current_term", "Term 1"),
        "current_year": data.get("current_year", datetime.now(timezone.utc).year),
        "onboarding_complete": True,
    }}, upsert=True)

    return {"message": "School setup complete"}


@router.post("/onboarding/setup")
async def onboarding_setup(data: dict, db=Depends(get_tenant_db)):
    """Legacy fresh-install setup: creates first admin account + saves school settings. No auth required."""
    from fastapi import HTTPException
    from fastapi.responses import JSONResponse
    user_count = await db.users.count_documents({})
    existing = await _get_settings(db)
    if user_count > 0 or (existing and existing.get("onboarding_complete")):
        raise HTTPException(status_code=400, detail="Setup has already been completed")

    name = data.get("admin_name", "").strip()
    email = data.get("admin_email", "").lower().strip()
    password = data.get("admin_password", "")
    if not name or not email or len(password) < 8:
        raise HTTPException(status_code=400, detail="Name, email and a password of at least 8 characters are required")

    user_id = f"user_{uuid.uuid4().hex[:12]}"
    await db.users.insert_one({
        "user_id": user_id, "email": email, "name": name,
        "picture": "", "role": "admin",
        "password_hash": _pwd.hash(password),
        "created_at": datetime.now(timezone.utc).isoformat(),
    })

    await db.school_settings.update_one({}, {"$set": {
        "school_name": data.get("school_name", "My School"),
        "school_type": data.get("school_type", "both"),
        "current_term": data.get("current_term", "Term 1"),
        "current_year": data.get("current_year", datetime.now(timezone.utc).year),
        "email_auth_enabled": True,
        "google_auth_enabled": data.get("google_auth_enabled", True),
        "onboarding_complete": True,
    }}, upsert=True)

    session_token = f"sess_{uuid.uuid4().hex}"
    await db.user_sessions.insert_one({
        "user_id": user_id, "session_token": session_token,
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat(),
    })

    resp = JSONResponse({"message": "Setup complete", "user_id": user_id})
    _set_session_cookie(resp, session_token)
    return resp


@router.post("/onboarding/complete")
async def complete_onboarding(data: dict, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    from fastapi import HTTPException
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    await db.school_settings.update_one(
        {},
        {"$set": {
            "school_name": data.get("school_name", "My School"),
            "school_type": data.get("school_type", "both"),
            "current_term": data.get("current_term", "Term 1"),
            "current_year": data.get("current_year", datetime.now(timezone.utc).year),
            "onboarding_complete": True,
        }},
        upsert=True
    )
    return {"message": "Onboarding complete"}


# ── Users ────────────────────────────────────────────────────────────────────

@router.get("/users")
async def get_users(user=Depends(get_current_user), db=Depends(get_tenant_db)):
    from fastapi import HTTPException
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return await db.users.find({}, {"_id": 0}).sort("name", 1).to_list(500)


@router.post("/users")
async def create_user(data: dict, request: Request, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    from fastapi import HTTPException
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    email = data.get("email", "").lower().strip()
    name = data.get("name", "")
    role = data.get("role", "teacher")
    _valid_roles = {"teacher", "wellbeing", "leadership", "admin", "screener", "professional"}
    if role not in _valid_roles:
        raise HTTPException(status_code=400, detail=f"Invalid role. Must be one of: {', '.join(sorted(_valid_roles))}")
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=409, detail="User with this email already exists")
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    new_user = {"user_id": user_id, "email": email, "name": name,
                "picture": "", "role": role,
                "created_at": datetime.now(timezone.utc).isoformat()}
    await db.users.insert_one({**new_user})
    await log_audit(db, user, "created", "user", user_id, f"{name} ({email})",
                    metadata={"role": role},
                    mirror_to_sa=True, request=request)
    return new_user


@router.post("/users/bulk")
async def bulk_create_users(data: dict, request: Request, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    """Bulk-create users from a parsed CSV. Rows with duplicate emails are skipped (not errored)."""
    from fastapi import HTTPException
    import re as _re
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    rows = data.get("users", [])
    if not rows:
        raise HTTPException(status_code=400, detail="No users provided")

    _valid_roles = {"teacher", "wellbeing", "leadership", "admin", "screener", "professional"}
    _email_re = _re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

    imported, skipped, errors = [], [], []
    # pre-load existing emails
    existing_emails = set()
    async for u in db.users.find({}, {"email": 1, "_id": 0}):
        if u.get("email"):
            existing_emails.add(u["email"].lower())

    seen_in_batch = set()
    for i, row in enumerate(rows):
        row_num = i + 2  # +1 for header, +1 for 1-indexed
        email = str(row.get("email") or row.get("Email") or "").lower().strip()
        name = str(row.get("name") or row.get("Name") or "").strip()
        role = str(row.get("role") or row.get("Role") or "teacher").lower().strip()

        if not email:
            errors.append({"row": row_num, "email": "", "error": "Email is required"})
            continue
        if not _email_re.match(email):
            errors.append({"row": row_num, "email": email, "error": "Invalid email format"})
            continue
        if role not in _valid_roles:
            errors.append({"row": row_num, "email": email, "error": f"Invalid role '{role}'"})
            continue
        if email in existing_emails or email in seen_in_batch:
            skipped.append({"row": row_num, "email": email, "reason": "already exists"})
            continue

        user_id = f"user_{uuid.uuid4().hex[:12]}"
        new_user = {"user_id": user_id, "email": email, "name": name,
                    "picture": "", "role": role,
                    "created_at": datetime.now(timezone.utc).isoformat()}
        await db.users.insert_one({**new_user})
        seen_in_batch.add(email)
        imported.append({"email": email, "name": name, "role": role})

    await log_audit(db, user, "bulk_import", "user", "", "Bulk user import",
                    bulk_count=len(imported), metadata={"skipped": len(skipped), "errors": len(errors)},
                    mirror_to_sa=True, request=request)

    return {"imported": len(imported), "skipped": len(skipped), "errors": errors, "skipped_list": skipped[:50]}


# ── Staff XLSX/CSV import ────────────────────────────────────────────────────
# PAYROLL_CLASS prefix → role map. Longer prefixes are matched first so "CES*"
# wins over "ES*". Every prefix ends with "*" in the user's spec but we strip it.
_STAFF_ROLE_RULES = [
    ("CES",   "screener"),     # Cover ES staff
    ("CLASS", "teacher"),      # Classroom teacher
    ("LEARN", "teacher"),      # Learning specialist
    ("LEAD",  "teacher"),      # Leading teacher
    ("PAR",   "teacher"),      # Paraprofessional
    ("AP",    "leadership"),   # Assistant principal
    ("PR",    "leadership"),   # Principal
    ("ES",    "screener"),     # Education support
]


def _role_for_payroll_class(payroll_class: str):
    """Return the mapped role (e.g. 'teacher') or None if no prefix matches.
    Longer prefixes checked first so 'CES' doesn't get mis-matched as 'ES'."""
    if not payroll_class:
        return None
    pc = payroll_class.strip().upper()
    for prefix, role in sorted(_STAFF_ROLE_RULES, key=lambda x: -len(x[0])):
        if pc.startswith(prefix):
            return role
    return None


# ── Shared: parse a CSV/XLSX upload into a list of dict rows ─────────────────
def _parse_spreadsheet(content: bytes, filename: str):
    """Parse CSV or XLSX bytes into a list of row dicts (keyed by header).
    Raises HTTPException(400) on unsupported format or empty file."""
    from fastapi import HTTPException
    import io
    import csv as _csv
    fname = (filename or "").lower()
    if fname.endswith(".csv"):
        reader = _csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        return [dict(r) for r in reader]
    if fname.endswith(".xlsx") or fname.endswith(".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise HTTPException(400, "File is empty.")
        header_idx = 0
        for i, r in enumerate(all_rows[:5]):
            if any(v not in (None, "") for v in r):
                header_idx = i
                break
        headers = [str(v or "").strip() for v in all_rows[header_idx]]
        out = []
        for r in all_rows[header_idx + 1:]:
            if not any(v not in (None, "") for v in r):
                continue
            row = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                v = r[i] if i < len(r) else None
                row[h] = "" if v is None else str(v).strip()
            out.append(row)
        return out
    raise HTTPException(400, "Unsupported file format. Please upload a CSV or XLSX file.")


def _row_get(row: dict, *keys: str) -> str:
    """Header-agnostic row accessor — matches regardless of case/underscores/spaces."""
    norm = lambda s: str(s).strip().upper().replace(" ", "_")
    wanted = {norm(k) for k in keys}
    for rk in row.keys():
        if norm(rk) in wanted:
            v = row[rk]
            if v is not None and str(v).strip() != "":
                return str(v).strip()
    return ""


def _detect_file_kind(rows: list) -> dict:
    """Sniff headers and return a hint about which upload this file is meant for.
    Used by preview endpoints so an accidental student → staff upload (or vice
    versa) is flagged BEFORE we touch the database."""
    if not rows:
        return {"looks_like": "unknown", "confidence": "low", "headers": []}
    headers = {str(h).strip().upper().replace(" ", "_") for h in rows[0].keys() if h}
    staff_markers = {"SFKEY", "E_MAIL", "PAYROLL_CLASS"}
    student_markers = {"STKEY", "SCHOOL_YEAR", "HOME_GROUP", "BIRTHDATE", "ENTRY", "PREF_NAME"}
    staff_hits = len(headers & staff_markers)
    student_hits = len(headers & student_markers)
    if staff_hits >= 2 and staff_hits > student_hits:
        return {"looks_like": "staff", "confidence": "high" if staff_hits >= 3 else "medium",
                "headers": sorted(headers)}
    if student_hits >= 2 and student_hits > staff_hits:
        return {"looks_like": "student", "confidence": "high" if student_hits >= 3 else "medium",
                "headers": sorted(headers)}
    return {"looks_like": "unknown", "confidence": "low", "headers": sorted(headers)}


def _analyse_staff_rows(rows: list, existing: dict) -> dict:
    """Dry-run analysis of a parsed staff spreadsheet.
    Returns counts + per-row decisions (add / update / skip / error)."""
    to_add, to_update, to_skip, errors, uncategorised = [], [], [], [], []
    seen_emails: set = set()
    for i, row in enumerate(rows):
        row_num = i + 2
        email = _row_get(row, "E_MAIL", "EMAIL", "email").lower()
        first = _row_get(row, "FIRST_NAME", "first_name")
        last = _row_get(row, "SURNAME", "last_name")
        sfkey = _row_get(row, "SFKEY", "staff_key")
        payroll_class = _row_get(row, "PAYROLL_CLASS", "payroll_class")

        # Mirror the importer: exclude Casual Relief Teachers.
        if payroll_class.upper().startswith("CRT"):
            to_skip.append({"row": row_num, "email": email, "reason": "CRT excluded"})
            continue

        if not email:
            errors.append({"row": row_num, "sfkey": sfkey, "error": "Missing E_MAIL"})
            continue
        if email in seen_emails:
            to_skip.append({"row": row_num, "email": email, "reason": "duplicate email in file"})
            continue
        seen_emails.add(email)

        name = f"{first} {last}".strip() or email
        role = _role_for_payroll_class(payroll_class)
        is_uncategorised = role is None
        if role is None:
            role = "teacher"
            uncategorised.append({"row": row_num, "email": email, "name": name,
                                  "payroll_class": payroll_class, "assigned": role})

        cur = existing.get(email)
        if cur:
            # Mirrors the import endpoint: existing users are preserved as-is
            # so manual role/name edits aren't clobbered on re-import.
            to_skip.append({"row": row_num, "email": email, "name": name,
                            "existing_role": cur.get("role"),
                            "reason": "already exists"})
        else:
            to_add.append({"row": row_num, "email": email, "name": name, "role": role,
                           "payroll_class": payroll_class, "uncategorised": is_uncategorised})
    return {
        "counts": {
            "add": len(to_add), "update": len(to_update),
            "skip": len(to_skip), "errors": len(errors),
            "uncategorised": len(uncategorised),
        },
        "add": to_add[:200], "update": to_update[:200],
        "skip": to_skip[:50], "errors": errors[:50],
        "uncategorised": uncategorised[:50],
    }


@router.post("/users/import-staff-preview")
async def import_staff_preview(
    file: UploadFile = File(...),
    user=Depends(get_current_user),
    db=Depends(get_tenant_db),
):
    """Dry-run: parse a staff file and report what WOULD happen on import.
    Nothing is written to the database. Used by the Settings UI to let admins
    verify the file contents — including that they're not accidentally
    uploading a student export — before committing."""
    from fastapi import HTTPException
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")

    content = await file.read()
    rows = _parse_spreadsheet(content, file.filename or "")
    kind = _detect_file_kind(rows)

    existing: dict = {}
    async for u in db.users.find({}, {"_id": 0, "user_id": 1, "email": 1, "role": 1, "name": 1, "sfkey": 1}):
        if u.get("email"):
            existing[u["email"].lower()] = u

    analysis = _analyse_staff_rows(rows, existing)
    return {
        "filename": file.filename,
        "total_rows": len(rows),
        "file_kind": kind,
        **analysis,
    }


@router.post("/users/import-staff")
async def import_staff(
    request: Request,
    file: UploadFile = File(...),
    user=Depends(get_current_user),
    db=Depends(get_tenant_db),
):
    """Import staff from an XLSX or CSV file using the SIS payroll export schema.

    Expected headers (row 1):
        SFKEY, FIRST_NAME, SURNAME, E_MAIL, PAYROLL_CLASS
    (STAFF_STATUS was previously required; rows are now always treated as
    Active per tenant request — the SIS export already excludes terminated
    staff.)

    Behaviour:
      • All rows are imported regardless of STAFF_STATUS (if present).
      • Role is derived from PAYROLL_CLASS prefix via _STAFF_ROLE_RULES.
      • Rows whose PAYROLL_CLASS doesn't match any rule default to 'teacher'
        and are surfaced in the response as `uncategorised` so the admin can
        reassign manually.
      • Matching: upsert by lowercase email.
    """
    from fastapi import HTTPException
    import io
    import csv as _csv
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")

    content = await file.read()
    fname = (file.filename or "").lower()

    if fname.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = _csv.DictReader(io.StringIO(text))
        rows = [dict(r) for r in reader]
    elif fname.endswith(".xlsx") or fname.endswith(".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise HTTPException(400, "File is empty.")
        header_idx = 0
        for i, r in enumerate(all_rows[:5]):
            if any(v not in (None, "") for v in r):
                header_idx = i
                break
        headers = [str(v or "").strip() for v in all_rows[header_idx]]
        rows = []
        for r in all_rows[header_idx + 1:]:
            if not any(v not in (None, "") for v in r):
                continue
            row = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                v = r[i] if i < len(r) else None
                row[h] = "" if v is None else str(v).strip()
            rows.append(row)
    else:
        raise HTTPException(400, "Unsupported file format. Please upload a CSV or XLSX file.")

    imported, updated, skipped, errors, uncategorised = [], [], [], [], []

    existing = {}
    async for u in db.users.find({}, {"_id": 0, "user_id": 1, "email": 1, "role": 1, "name": 1}):
        if u.get("email"):
            existing[u["email"].lower()] = u

    now_iso = datetime.now(timezone.utc).isoformat()

    for i, row in enumerate(rows):
        row_num = i + 2
        def _g(*keys, r=row):
            for k in keys:
                for rk in r.keys():
                    if rk.strip().upper().replace(" ", "_") == k.upper().replace(" ", "_"):
                        v = r[rk]
                        if v is not None and str(v).strip() != "":
                            return str(v).strip()
            return ""

        status = _g("STAFF_STATUS", "status").lower()
        # Per tenant request (2026-02): STAFF_STATUS is no longer required —
        # all rows are assumed Active. We keep the read only so a future
        # re-introduction of the column is backwards-compatible.
        _ = status

        email = _g("E_MAIL", "EMAIL", "email").lower()
        first = _g("FIRST_NAME", "first_name")
        last = _g("SURNAME", "last_name")
        sfkey = _g("SFKEY", "staff_key")
        payroll_class = _g("PAYROLL_CLASS", "payroll_class")

        # Skip Casual Relief Teachers — tenant-requested exclusion because
        # CRTs are short-term placements and shouldn't count as permanent
        # staff for wellbeing/MTSS purposes.
        if payroll_class.upper().startswith("CRT"):
            skipped.append({"row": row_num, "email": email, "reason": "CRT excluded"})
            continue

        if not email:
            errors.append({"row": row_num, "sfkey": sfkey, "error": "Missing E_MAIL"})
            continue

        name = f"{first} {last}".strip() or email
        role = _role_for_payroll_class(payroll_class)
        if role is None:
            role = "teacher"
            uncategorised.append({"row": row_num, "email": email, "name": name,
                                  "payroll_class": payroll_class, "assigned": role})

        if email in existing:
            # Per tenant request (2026-02): once a staff user exists, do NOT
            # overwrite them on re-import — this preserves manual edits such
            # as role changes, professional settings, or name corrections
            # the admin has made in User Management.
            skipped.append({"row": row_num, "email": email, "reason": "already exists"})
        else:
            new_user = {
                "user_id": f"user_{uuid.uuid4().hex[:12]}",
                "email": email,
                "name": name,
                "picture": "",
                "role": role,
                "sfkey": sfkey or None,
                "created_at": now_iso,
            }
            new_user = {k: v for k, v in new_user.items() if v is not None}
            await db.users.insert_one({**new_user})
            existing[email] = new_user
            imported.append({"email": email, "name": name, "role": role})

    await log_audit(
        db, user, "bulk_import", "user", "",
        f"Staff import — {file.filename}",
        bulk_count=len(imported) + len(updated),
        metadata={
            "imported": len(imported),
            "updated": len(updated),
            "skipped": len(skipped),
            "errors": len(errors),
            "uncategorised": len(uncategorised),
        },
        mirror_to_sa=True, request=request,
    )

    return {
        "imported": len(imported),
        "updated": len(updated),
        "skipped": len(skipped),
        "errors": errors,
        "uncategorised": uncategorised[:50],
        "total": len(rows),
    }


@router.get("/users/bulk-template")
async def users_bulk_template(user=Depends(get_current_user)):
    """Download a CSV template for bulk user upload."""
    from fastapi import HTTPException
    from fastapi.responses import Response
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    csv_text = (
        "email,name,role\n"
        "jane.smith@school.edu.au,Jane Smith,teacher\n"
        "john.doe@school.edu.au,John Doe,wellbeing\n"
        "principal@school.edu.au,Alex Principal,admin\n"
        "# Valid roles: teacher | wellbeing | leadership | admin | screener | professional\n"
    )
    return Response(
        content=csv_text,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=welltrack-users-template.csv"},
    )


@router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, data: dict, request: Request, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    from fastapi import HTTPException
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    role = data.get("role")
    if role not in ["teacher", "wellbeing", "leadership", "admin", "screener", "professional"]:
        raise HTTPException(status_code=400, detail="Invalid role")
    await db.users.update_one({"user_id": user_id}, {"$set": {"role": role}})
    target = await db.users.find_one({"user_id": user_id}, {"name": 1, "email": 1, "_id": 0})
    await log_audit(db, user, "updated", "user", user_id,
                    f"{target.get('name','?')} ({target.get('email','?')})" if target else user_id,
                    changes={"role": {"new": role}},
                    mirror_to_sa=True, request=request)
    return {"message": "Role updated"}


@router.delete("/users/{user_id}")
async def delete_user(user_id: str, request: Request, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    from fastapi import HTTPException
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    if user_id == user["user_id"]:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    target = await db.users.find_one({"user_id": user_id}, {"name": 1, "email": 1, "_id": 0})
    await db.users.delete_one({"user_id": user_id})
    await db.user_sessions.delete_many({"user_id": user_id})
    await log_audit(db, user, "deleted", "user", user_id,
                    f"{target.get('name','?')} ({target.get('email','?')})" if target else user_id,
                    mirror_to_sa=True, request=request)
    return {"message": "User deleted"}


@router.put("/users/{user_id}")
async def update_user(user_id: str, data: dict, request: Request, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    from fastapi import HTTPException
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")

    # Get target user
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    # Fields that can be updated
    allowed = {"name", "email", "picture", "role"}
    update = {k: v for k, v in data.items() if k in allowed}

    if not update:
        raise HTTPException(status_code=400, detail="No valid fields to update")

    # Validate email uniqueness if changing email
    if "email" in update and update["email"] != target.get("email"):
        existing = await db.users.find_one({"email": update["email"]})
        if existing:
            raise HTTPException(status_code=409, detail="A user with this email already exists")
        update["email"] = update["email"].lower().strip()

    # Validate role if changing
    if "role" in update and update["role"] not in ["teacher", "wellbeing", "leadership", "admin", "screener", "professional"]:
        raise HTTPException(status_code=400, detail="Invalid role")

    await db.users.update_one({"user_id": user_id}, {"$set": update})
    updated = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})

    await log_audit(db, user, "updated", "user", user_id,
                    f"{updated.get('name','?')} ({updated.get('email','?')})" if updated else user_id,
                    changes={k: v for k, v in update.items()},
                    mirror_to_sa=True, request=request)

    return updated


@router.put("/users/{user_id}/professional")
async def update_user_professional(user_id: str, data: dict, request: Request, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    from fastapi import HTTPException
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    allowed = {
        "professional_type", "appointment_access",
        "visit_days", "accessible_intervention_types", "cross_professional_view",
    }
    update = {k: v for k, v in data.items() if k in allowed}
    if not update:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    await db.users.update_one({"user_id": user_id}, {"$set": update})
    updated = await db.users.find_one({"user_id": user_id}, {"_id": 0, "hashed_password": 0, "password_hash": 0})
    await log_audit(db, user, "updated", "user", user_id,
                    f"{updated.get('name','?')} — professional settings" if updated else user_id,
                    changes={k: v for k, v in update.items()},
                    mirror_to_sa=True, request=request)
    return updated
