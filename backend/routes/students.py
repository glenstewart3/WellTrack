from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from typing import Optional
import uuid, zipfile, io, re, os, csv
from pathlib import Path
from PIL import Image

from deps import get_tenant_db
import asyncio
from helpers import get_current_user, get_student_attendance_pct, compute_mtss_tier, \
    get_bulk_attendance_stats, get_latest_saebrs_bulk, get_latest_saebrs_plus_bulk
from models import Student
from utils.audit import log_audit

# Base uploads directory
_UPLOADS_BASE = Path(os.environ.get("UPLOADS_DIR", str(Path(__file__).resolve().parent.parent / "uploads")))


def _get_photos_dir(request: Request) -> Path:
    """Return tenant-scoped photos directory. Creates it if missing."""
    slug = getattr(request.state, "tenant_slug", None) or "default"
    path = _UPLOADS_BASE / slug / "student_photos"
    path.mkdir(parents=True, exist_ok=True)
    return path


router = APIRouter()


@router.get("/students")
async def get_students(class_name: Optional[str] = None, year_level: Optional[str] = None,
                       status: Optional[str] = None,
                       user=Depends(get_current_user), db=Depends(get_tenant_db)):
    query = {"enrolment_status": status if status in ("active", "archived") else "active"}
    if class_name:
        query["class_name"] = class_name
    if year_level:
        query["year_level"] = year_level
    return await db.students.find(query, {"_id": 0}).sort("last_name", 1).to_list(500)


@router.post("/students")
async def create_student(student: Student, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    d = student.model_dump()
    await db.students.insert_one({**d})
    await log_audit(db, user, "created", "student", d.get("student_id", ""),
                    f"{d.get('first_name', '')} {d.get('last_name', '')}".strip())
    return d


@router.post("/students/import")
async def import_students(data: dict, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    """Import students from a CSV/XLSX export.

    Supports two formats:
      • NEW (2026+): STKEY, FIRST_NAME, PREF_NAME, SURNAME, FAMILY, GENDER, BIRTHDATE,
        ENTRY, HOME_GROUP, SCHOOL_YEAR, KOORIE
      • LEGACY:     Import Identifier, First Name, Preferred Name, Surname, Year
        Level, Form Group, Base Role, User Status, ...
    """
    rows = data.get("students", [])
    if not rows:
        raise HTTPException(status_code=400, detail="No student data provided")

    def _get(row, *keys):
        for k in keys:
            v = row.get(k)
            if v is not None and str(v).strip() != "":
                return str(v).strip()
        return ""

    def _date_iso(v):
        if not v:
            return ""
        s = str(v).strip()
        # Excel often exports "2024-01-28 00:00:00"
        if " " in s:
            s = s.split(" ", 1)[0]
        return s[:10]

    _koorie_map = {
        "N": "No", "K": "Koorie", "B": "Both (ATSI)",
        "T": "Torres Strait Islander", "A": "Aboriginal",
    }

    imported, updated, errors = [], [], []
    for i, row in enumerate(rows):
        # Legacy-only filters
        base_role = _get(row, "Base Role", "base_role") or "Student"
        user_status = _get(row, "User Status", "user_status") or "Active"
        if base_role.lower() not in ("student", "") or user_status.lower() == "inactive":
            continue

        stkey = _get(row, "STKEY", "Import Identifier", "SussiId", "sussi_id")
        first_name     = _get(row, "FIRST_NAME", "First Name", "first_name")
        preferred_name = _get(row, "PREF_NAME",  "Preferred Name", "preferred_name")
        last_name      = _get(row, "SURNAME",    "Surname", "last_name")
        family         = _get(row, "FAMILY",     "family")
        gender         = _get(row, "GENDER",     "gender")
        dob            = _date_iso(_get(row, "BIRTHDATE", "date_of_birth", "dob"))
        entry_date     = _date_iso(_get(row, "ENTRY", "enrolment_date", "entry_date"))
        class_name     = _get(row, "HOME_GROUP", "Form Group", "class_name")
        year_level     = _get(row, "SCHOOL_YEAR","Year Level", "year_level")
        koorie_raw     = _get(row, "KOORIE",     "koorie").upper()
        koorie         = _koorie_map.get(koorie_raw, koorie_raw) if koorie_raw else ""
        # Normalise SCHOOL_YEAR:
        #   "0"/"00"        → "Foundation"
        #   "01".."12"      → "Year 01".."Year 12" (keep zero-padding)
        #   "1".."12"       → "Year 1".."Year 12"
        #   "Prep"/"F"/"K"  → "Foundation"
        #   anything else   → passed through unchanged
        if year_level:
            yl = year_level.strip()
            if yl in ("0", "00") or yl.lower() in ("prep", "f", "k", "kinder", "foundation"):
                year_level = "Foundation"
            elif yl.isdigit() and 1 <= int(yl) <= 12:
                year_level = f"Year {yl}"
            # else keep original (e.g. "Year 1" already, "VCE", etc.)

        if not stkey and not first_name and not last_name:
            errors.append({"row": i + 1, "error": "Missing student identifier"})
            continue

        student_doc = {
            "first_name": first_name or stkey,
            "preferred_name": preferred_name or None,
            "last_name": last_name,
            "year_level": year_level,
            "class_name": class_name,
            "gender": gender,
            "date_of_birth": dob,
            "family": family or None,
            "entry_date": entry_date or None,
            "koorie": koorie or None,
            "enrolment_status": "active",
            "sussi_id": stkey, "external_id": stkey,
        }
        # Drop empty optional keys so partial updates don't overwrite existing data
        student_doc = {k: v for k, v in student_doc.items() if v != "" and v is not None}

        if stkey:
            existing = await db.students.find_one({"sussi_id": stkey})
            if existing:
                await db.students.update_one({"sussi_id": stkey}, {"$set": student_doc})
                updated.append(stkey)
            else:
                student_doc["student_id"] = f"stu_{uuid.uuid4().hex[:8]}"
                await db.students.insert_one({**student_doc})
                imported.append(student_doc)
        else:
            student_doc["student_id"] = f"stu_{uuid.uuid4().hex[:8]}"
            await db.students.insert_one({**student_doc})
            imported.append(student_doc)

    await log_audit(db, user, "bulk_import", "student", "", "Student bulk import",
                    bulk_count=len(imported) + len(updated),
                    metadata={"imported": len(imported), "updated": len(updated), "errors": len(errors)})
    return {"imported": len(imported), "updated": len(updated), "errors": errors, "total": len(rows)}


@router.post("/students/import-file")
async def import_students_from_file(file: UploadFile = File(...), user=Depends(get_current_user), db=Depends(get_tenant_db)):
    """Accept a CSV or XLSX file directly (multipart upload), parse it into rows,
    and hand the rows off to the same processing pipeline used by /students/import."""
    if user.get("role") not in ["admin", "leadership"]:
        raise HTTPException(403, "Admin or leadership access required")

    content = await file.read()
    fname = (file.filename or "").lower()

    if fname.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = [dict(r) for r in reader]
    elif fname.endswith(".xlsx") or fname.endswith(".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise HTTPException(400, "File is empty.")
        # First non-empty row = headers
        header_idx = 0
        for i, r in enumerate(all_rows[:5]):
            if any(v not in (None, "") for v in r):
                header_idx = i
                break
        headers = [str(v or "").strip() for v in all_rows[header_idx]]
        rows = []
        for r in all_rows[header_idx + 1:]:
            if not any(v not in (None, "") for v in r):
                continue
            row = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                v = r[i] if i < len(r) else None
                row[h] = "" if v is None else str(v).strip()
            rows.append(row)
    else:
        raise HTTPException(400, "Unsupported file format. Please upload a CSV or XLSX file.")

    # Delegate to the same processing logic by calling the JSON endpoint handler
    return await import_students({"students": rows}, user=user, db=db)


@router.get("/students/summary")
async def get_students_summary(class_name: Optional[str] = None, year_level: Optional[str] = None,
                                status: Optional[str] = None,
                                user=Depends(get_current_user), db=Depends(get_tenant_db)):
    query = {"enrolment_status": status if status in ("active", "archived") else "active"}
    if class_name:
        query["class_name"] = class_name
    if year_level:
        query["year_level"] = year_level
    students = await db.students.find(query, {"_id": 0}).sort("last_name", 1).to_list(500)
    student_ids = [s["student_id"] for s in students]

    saebrs_map, plus_map, att_map, active_int_docs = await asyncio.gather(
        get_latest_saebrs_bulk(db, student_ids),
        get_latest_saebrs_plus_bulk(db, student_ids),
        get_bulk_attendance_stats(db, student_ids),
        db.interventions.find({"student_id": {"$in": student_ids}, "status": "active"},
                              {"_id": 0, "student_id": 1}).to_list(1000),
    )
    active_int_count: dict = {}
    for intv in active_int_docs:
        sid = intv["student_id"]
        active_int_count[sid] = active_int_count.get(sid, 0) + 1

    result = []
    for s in students:
        sid = s["student_id"]
        saebrs = saebrs_map.get(sid)
        plus = plus_map.get(sid)
        att_pct = att_map.get(sid, {}).get("pct", 100.0)

        if saebrs and plus:
            tier = compute_mtss_tier(saebrs["risk_level"], plus["wellbeing_tier"], att_pct)
        elif saebrs:
            tier = 3 if saebrs["risk_level"] == "High Risk" else (2 if saebrs["risk_level"] == "Some Risk" else 1)
        else:
            tier = None

        result.append({
            **s,
            "mtss_tier": tier,
            "saebrs_risk": saebrs["risk_level"] if saebrs else "Not Screened",
            "saebrs_total": saebrs["total_score"] if saebrs else None,
            "wellbeing_tier": plus["wellbeing_tier"] if plus else None,
            "wellbeing_total": plus["wellbeing_total"] if plus else None,
            "attendance_pct": round(att_pct, 1),
            "active_interventions": active_int_count.get(sid, 0),
        })
    return result


@router.get("/students/{student_id}")
async def get_student(student_id: str, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    s = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Student not found")
    return s


@router.get("/students/{student_id}/profile")
async def get_student_profile(student_id: str, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    saebrs_results = await db.saebrs_results.find({"student_id": student_id}, {"_id": 0}).sort("created_at", 1).to_list(20)
    saebrs_plus = await db.self_report_results.find({"student_id": student_id}, {"_id": 0}).sort("created_at", 1).to_list(20)
    interventions = await db.interventions.find({"student_id": student_id}, {"_id": 0}).sort("created_at", -1).to_list(20)
    case_notes = await db.case_notes.find({"student_id": student_id}, {"_id": 0}).sort("date", -1).to_list(20)
    att_pct = await get_student_attendance_pct(db, student_id)
    attendance_records = await db.attendance_records.find({"student_id": student_id}, {"_id": 0}).sort("date", -1).to_list(50)
    alerts = await db.alerts.find({"student_id": student_id, "resolved": False}, {"_id": 0}).to_list(10)

    latest_saebrs = saebrs_results[-1] if saebrs_results else None
    latest_plus = saebrs_plus[-1] if saebrs_plus else None

    if latest_saebrs and latest_plus:
        tier = compute_mtss_tier(latest_saebrs["risk_level"], latest_plus["wellbeing_tier"], att_pct)
    elif latest_saebrs:
        tier = 3 if latest_saebrs["risk_level"] == "High Risk" else (2 if latest_saebrs["risk_level"] == "Some Risk" else 1)
    else:
        tier = None

    return {
        "student": student, "mtss_tier": tier,
        "attendance_pct": round(att_pct, 1),
        "saebrs_results": saebrs_results,
        "self_report_results": saebrs_plus,
        "interventions": interventions,
        "case_notes": case_notes,
        "attendance_records": attendance_records,
        "alerts": alerts
    }


@router.put("/students/{student_id}/external-id")
async def set_student_external_id(student_id: str, data: dict, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    if user.get("role") not in ["admin", "leadership"]:
        raise HTTPException(403, "Access denied")
    ext_id = data.get("external_id", "").strip().upper()
    await db.students.update_one({"student_id": student_id}, {"$set": {"external_id": ext_id}})
    await log_audit(db, user, "updated", "student", student_id, f"External ID → {ext_id}",
                    changes={"external_id": {"new": ext_id}})
    return {"message": "External ID updated", "external_id": ext_id}


@router.put("/students/bulk-archive")
async def bulk_archive_students(data: dict, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    if user.get("role") not in ["admin", "leadership"]:
        raise HTTPException(403, "Access denied")
    ids = data.get("student_ids", [])
    if not ids:
        raise HTTPException(400, "No student IDs provided")
    result = await db.students.update_many(
        {"student_id": {"$in": ids}},
        {"$set": {"enrolment_status": "archived"}}
    )
    await log_audit(db, user, "bulk_archive", "student", "", "Bulk archive students", bulk_count=result.modified_count)
    return {"archived": result.modified_count}


@router.put("/students/bulk-reactivate")
async def bulk_reactivate_students(data: dict, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    if user.get("role") not in ["admin", "leadership"]:
        raise HTTPException(403, "Access denied")
    ids = data.get("student_ids", [])
    if not ids:
        raise HTTPException(400, "No student IDs provided")
    result = await db.students.update_many(
        {"student_id": {"$in": ids}},
        {"$set": {"enrolment_status": "active"}}
    )
    await log_audit(db, user, "bulk_reactivate", "student", "", "Bulk reactivate students", bulk_count=result.modified_count)
    return {"reactivated": result.modified_count}


@router.put("/students/{student_id}")
async def update_student(student_id: str, data: dict, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    if user.get("role") not in ["admin", "leadership"]:
        raise HTTPException(403, "Access denied")
    allowed = ["first_name", "last_name", "preferred_name", "year_level", "class_name", "teacher", "gender", "date_of_birth"]
    update_data = {k: v for k, v in data.items() if k in allowed}
    if not update_data:
        raise HTTPException(400, "No valid fields to update")
    result = await db.students.update_one({"student_id": student_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(404, "Student not found")
    updated = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    await log_audit(db, user, "updated", "student", student_id,
                    f"{updated.get('first_name','')} {updated.get('last_name','')}".strip(),
                    changes={k: v for k, v in update_data.items()})
    return updated


@router.post("/students/{student_id}/photo")
async def upload_single_student_photo(request: Request, student_id: str, file: UploadFile = File(...), user=Depends(get_current_user), db=Depends(get_tenant_db)):
    if user.get("role") not in ["admin", "leadership"]:
        raise HTTPException(403, "Access denied")
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0, "student_id": 1})
    if student is None:
        raise HTTPException(404, "Student not found")
    photos_dir = _get_photos_dir(request)
    slug = getattr(request.state, "tenant_slug", None) or "default"
    content = await file.read()
    img = Image.open(io.BytesIO(content))
    img = img.convert("RGB")
    img.thumbnail((400, 400), Image.LANCZOS)
    photo_filename = f"{student_id}.jpg"
    photo_path = photos_dir / photo_filename
    img.save(str(photo_path), "JPEG", quality=82, optimize=True)
    photo_url = f"/api/student-photos/{slug}/{photo_filename}"
    await db.students.update_one({"student_id": student_id}, {"$set": {"photo_url": photo_url}})
    await log_audit(db, user, "uploaded", "photo", student_id, f"Photo uploaded for student {student_id}")
    return {"photo_url": photo_url}


@router.delete("/students/{student_id}/photo")
async def remove_student_photo(request: Request, student_id: str, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    if user.get("role") not in ["admin", "leadership"]:
        raise HTTPException(403, "Access denied")
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0, "student_id": 1, "photo_url": 1})
    if student is None:
        raise HTTPException(404, "Student not found")
    photo_url = student.get("photo_url")
    if photo_url:
        photos_dir = _get_photos_dir(request)
        photo_path = photos_dir / Path(photo_url).name
        if photo_path.exists():
            photo_path.unlink()
    await db.students.update_one({"student_id": student_id}, {"$unset": {"photo_url": ""}})
    await log_audit(db, user, "deleted", "photo", student_id, f"Photo removed for student {student_id}")
    return {"message": "Photo removed"}


@router.post("/students/refresh-photos")
async def refresh_student_photos(request: Request, user=Depends(get_current_user), db=Depends(get_tenant_db)):
    """Reconcile student photo_urls with files actually on disk.

    - Clears photo_url on any student whose referenced file is missing (prevents
      broken-image icons in the UI when the uploads directory has been reset).
    - Re-links any on-disk `stu_xxx.jpg` file back to its student record if the
      student exists but has no photo_url set.

    Returns counts for audit clarity.
    """
    if user.get("role") not in ["admin", "leadership"]:
        raise HTTPException(403, "Admin or leadership access required")

    slug = getattr(request.state, "tenant_slug", None) or "default"
    photos_dir = _UPLOADS_BASE / slug / "student_photos"
    photos_dir.mkdir(parents=True, exist_ok=True)

    # What's on disk?
    on_disk = {f.stem: f.name for f in photos_dir.iterdir() if f.is_file()}

    cleared = 0
    relinked = 0

    # 1. Clear stale photo_urls (file missing)
    async for s in db.students.find(
        {"photo_url": {"$exists": True, "$ne": None, "$ne": ""}},
        {"_id": 0, "student_id": 1, "photo_url": 1},
    ):
        url = s.get("photo_url") or ""
        fname = url.rsplit("/", 1)[-1]
        stem = fname.rsplit(".", 1)[0]
        if stem not in on_disk:
            await db.students.update_one(
                {"student_id": s["student_id"]},
                {"$unset": {"photo_url": ""}},
            )
            cleared += 1

    # 2. Re-link orphan files whose stem matches a student without a photo_url
    for stu_id, fname in on_disk.items():
        stu = await db.students.find_one(
            {"student_id": stu_id},
            {"_id": 0, "photo_url": 1},
        )
        if stu and not stu.get("photo_url"):
            photo_url = f"/api/student-photos/{slug}/{fname}"
            await db.students.update_one(
                {"student_id": stu_id},
                {"$set": {"photo_url": photo_url}},
            )
            relinked += 1

    await log_audit(db, user, "updated", "photo", "", "Photo URLs refreshed",
                    metadata={"cleared": cleared, "relinked": relinked,
                              "files_on_disk": len(on_disk)})
    return {
        "cleared_stale": cleared,
        "relinked_orphans": relinked,
        "files_on_disk": len(on_disk),
    }


@router.post("/students/upload-photos")
async def upload_student_photos(request: Request, file: UploadFile = File(...), user=Depends(get_current_user), db=Depends(get_tenant_db)):
    if user.get("role") not in ["admin", "leadership"]:
        raise HTTPException(403, "Access denied")

    fname_lower = (file.filename or "").lower()
    if not fname_lower.endswith(".zip"):
        raise HTTPException(400, "Please upload a ZIP file")

    photos_dir = _get_photos_dir(request)
    slug = getattr(request.state, "tenant_slug", None) or "default"
    staff_photos_dir = _UPLOADS_BASE / slug / "staff_photos"
    staff_photos_dir.mkdir(parents=True, exist_ok=True)
    content = await file.read()

    matched, unmatched = [], []
    matched_staff, unmatched_staff = [], []

    # ── Name normalisation + index-once strategy ─────────────────────────────
    # Covers: curly apostrophes, straight apostrophes, hyphens, spaces, accents,
    # case variance, compound last names (ERSCH- MAHY), multi-token first names
    # (Nur Fathiya Adira), nicknames (Michael → Mike), and filename formats
    # both "Last, First.jpg" and "First Last.jpg" / "Last_First.jpg".
    import unicodedata
    from difflib import SequenceMatcher

    def _norm(s: str) -> str:
        if not s:
            return ""
        s = unicodedata.normalize("NFKD", str(s))
        s = "".join(c for c in s if not unicodedata.combining(c))
        return "".join(c.lower() for c in s if c.isalnum())

    def _tokens(s: str):
        """Return lowercase alphanumeric tokens, splitting on any non-alnum."""
        if not s:
            return []
        s = unicodedata.normalize("NFKD", str(s))
        s = "".join(c for c in s if not unicodedata.combining(c))
        buf, out = [], []
        for c in s:
            if c.isalnum():
                buf.append(c.lower())
            elif buf:
                out.append("".join(buf))
                buf = []
        if buf:
            out.append("".join(buf))
        return out

    def _first_name_similarity(a: str, b: str) -> float:
        """Return a similarity score in [0,1] between two normalised first names.
        Handles nicknames via prefix detection and character-level similarity."""
        if not a or not b:
            return 0.0
        if a == b:
            return 1.0
        # Nickname heuristic: one is a >=3-char prefix of the other
        shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
        if len(shorter) >= 3 and longer.startswith(shorter):
            return 0.9
        # Shared prefix heuristic (Mike/Michael share "mi"; Kate/Katherine share "kat")
        # Give a boost when ≥3 initial chars match
        common_prefix = 0
        for i in range(min(len(shorter), len(longer))):
            if shorter[i] == longer[i]:
                common_prefix += 1
            else:
                break
        base = SequenceMatcher(None, a, b).ratio()
        if common_prefix >= 3:
            return max(base, 0.75)
        if common_prefix >= 2 and len(shorter) >= 3:
            return max(base, 0.6)
        return base

    # Index active students once
    students_raw = await db.students.find(
        {"enrolment_status": "active"},
        {"_id": 0, "student_id": 1, "first_name": 1, "last_name": 1, "preferred_name": 1},
    ).to_list(5000)

    index = []
    for s in students_raw:
        fn, ln, pn = s.get("first_name", ""), s.get("last_name", ""), s.get("preferred_name", "")
        index.append({
            "student_id": s["student_id"],
            "fn_norm": _norm(fn),
            "ln_norm": _norm(ln),
            "pn_norm": _norm(pn),
            "fn_tokens": _tokens(fn),
            "ln_tokens": _tokens(ln),
        })

    # Index staff users — split "name" into first + last tokens
    staff_raw = await db.users.find(
        {},
        {"_id": 0, "user_id": 1, "name": 1, "email": 1},
    ).to_list(2000)

    staff_index = []
    for u in staff_raw:
        full = (u.get("name") or "").strip()
        if not full:
            continue
        toks = _tokens(full)
        if not toks:
            continue
        # Assume "First [Middle] Last" — last token = surname, rest = first/given
        staff_index.append({
            "user_id": u["user_id"],
            "email": u.get("email", ""),
            "name": full,
            "ln_norm": _norm(toks[-1]),
            "fn_norm": _norm(toks[0]),
            "all_tokens": toks,
        })

    def _match_student_pair(last_name: str, first_name: str):
        """Try to match one (last, first) ordering. Returns rec or None."""
        ln_n = _norm(last_name)
        fn_n = _norm(first_name)
        if not ln_n or not fn_n:
            return None

        # 1. Exact normalised match on last + (first OR preferred)
        for rec in index:
            if rec["ln_norm"] == ln_n and (rec["fn_norm"] == fn_n or rec["pn_norm"] == fn_n):
                return rec
        # 2. Last-name token match + exact or preferred first name
        for rec in index:
            if ln_n in rec["ln_tokens"] and (rec["fn_norm"] == fn_n or rec["pn_norm"] == fn_n):
                return rec
        # 3. Exact last + first-name token match (multi-token first names)
        for rec in index:
            if rec["ln_norm"] == ln_n and fn_n in rec["fn_tokens"]:
                return rec
        # 4. Both sides token-level
        for rec in index:
            if ln_n in rec["ln_tokens"] and fn_n in rec["fn_tokens"]:
                return rec
        # 5. Last-name match + fuzzy/nickname first-name (pick best above threshold)
        candidates = []
        for rec in index:
            if rec["ln_norm"] == ln_n or ln_n in rec["ln_tokens"]:
                scores = [_first_name_similarity(fn_n, rec["fn_norm"])]
                if rec["pn_norm"]:
                    scores.append(_first_name_similarity(fn_n, rec["pn_norm"]))
                for t in rec["fn_tokens"]:
                    scores.append(_first_name_similarity(fn_n, t))
                best = max(scores) if scores else 0.0
                candidates.append((best, rec))
        if candidates:
            candidates.sort(key=lambda x: x[0], reverse=True)
            top_score, top_rec = candidates[0]
            # 5a. Strong similarity → use it
            if top_score >= 0.6:
                return top_rec
            # 5b. Unique last-name match → match regardless of first name
            # (covers nickname-only-in-photo and typos)
            if len(candidates) == 1:
                return top_rec
        return None

    def _match_student(last_name: str, first_name: str):
        # Try given order first, then swap (handles "First Last" filename convention)
        rec = _match_student_pair(last_name, first_name)
        if rec:
            return rec
        return _match_student_pair(first_name, last_name)

    def _match_staff_pair(last_name: str, first_name: str):
        ln_n = _norm(last_name)
        fn_n = _norm(first_name)
        if not ln_n or not fn_n:
            return None
        # 1. Exact: last-token matches staff surname AND first-token matches staff given name
        for rec in staff_index:
            if rec["ln_norm"] == ln_n and rec["fn_norm"] == fn_n:
                return rec
        # 2. Both appear as tokens anywhere in the staff name
        for rec in staff_index:
            if ln_n in rec["all_tokens"] and fn_n in rec["all_tokens"]:
                return rec
        # 3. Strict last-name match + best-scoring fuzzy first name (nicknames)
        candidates = []
        for rec in staff_index:
            if rec["ln_norm"] == ln_n or ln_n in rec["all_tokens"]:
                scores = [_first_name_similarity(fn_n, rec["fn_norm"])]
                for t in rec["all_tokens"][:-1]:
                    scores.append(_first_name_similarity(fn_n, t))
                best = max(scores) if scores else 0.0
                candidates.append((best, rec))
        if candidates:
            candidates.sort(key=lambda x: x[0], reverse=True)
            top_score, top_rec = candidates[0]
            # 3a. Any reasonable similarity → use it (catches nicknames)
            if top_score >= 0.5:
                return top_rec
            # 3b. Only one staff member with that surname → match them anyway
            if len(candidates) == 1:
                return top_rec
        return None

    def _match_staff(last_name: str, first_name: str):
        # Try given order, then swap for "First Last" filename convention
        rec = _match_staff_pair(last_name, first_name)
        if rec:
            return rec
        return _match_staff_pair(first_name, last_name)

    def _parse_stem(stem: str):
        """Parse a photo filename stem into (last, first) candidates.
        Returns (last_guess, first_guess) — caller then calls the matcher which
        also tries the swapped order. Supports 'Last, First', 'Last_First',
        'Last-First' (when explicit), and falls back to the last-space-separated
        token being treated as the surname ('First Middle Last' convention)."""
        stem = (stem or "").strip()
        if not stem:
            return None, None
        # Prefer comma if present
        if "," in stem:
            last_raw, first_raw = stem.split(",", 1)
            return last_raw.strip(), first_raw.strip()
        # Underscore as separator (common school export: "Smith_Jane.jpg")
        if "_" in stem and " " not in stem:
            parts = stem.split("_", 1)
            if len(parts) == 2:
                return parts[0].strip(), parts[1].strip()
        # Fallback: space-separated, last word = surname ("Jane Smith.jpg")
        toks = stem.split()
        if len(toks) >= 2:
            return toks[-1].strip(), " ".join(toks[:-1]).strip()
        return None, None

    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue

            path_parts = Path(info.filename).parts
            is_staff_photo = any(p.lower().startswith("staff") for p in path_parts)

            stem_raw = Path(info.filename).stem
            ext = Path(info.filename).suffix.lower()
            if ext not in (".jpg", ".jpeg", ".png"):
                continue

            stem = stem_raw.strip()
            last_name, first_name = _parse_stem(stem)

            if not last_name or not first_name:
                (unmatched_staff if is_staff_photo else unmatched).append(stem_raw)
                continue

            # ── Staff photo: match to user account ──────────────────────────────
            if is_staff_photo:
                staff_rec = _match_staff(last_name, first_name)
                if not staff_rec:
                    unmatched_staff.append(f"{last_name}, {first_name}")
                    continue
                user_id = staff_rec["user_id"]
                photo_filename = f"{user_id}.jpg"
                photo_path = staff_photos_dir / photo_filename
                with zf.open(info) as img_bytes:
                    img = Image.open(img_bytes)
                    img = img.convert("RGB")
                    img.thumbnail((400, 400), Image.LANCZOS)
                    img.save(str(photo_path), "JPEG", quality=82, optimize=True)
                photo_url = f"/api/staff-photos/{slug}/{photo_filename}"
                await db.users.update_one(
                    {"user_id": user_id},
                    {"$set": {"picture": photo_url}},
                )
                matched_staff.append(f"{last_name}, {first_name}")
                continue

            # ── Student photo ───────────────────────────────────────────────────
            rec = _match_student(last_name, first_name)

            if not rec:
                unmatched.append(f"{last_name}, {first_name}")
                continue

            student_id = rec["student_id"]
            photo_filename = f"{student_id}.jpg"
            photo_path = photos_dir / photo_filename

            with zf.open(info) as img_bytes:
                img = Image.open(img_bytes)
                img = img.convert("RGB")
                img.thumbnail((400, 400), Image.LANCZOS)
                img.save(str(photo_path), "JPEG", quality=82, optimize=True)

            photo_url = f"/api/student-photos/{slug}/{photo_filename}"
            await db.students.update_one(
                {"student_id": student_id},
                {"$set": {"photo_url": photo_url}},
            )
            matched.append(f"{last_name}, {first_name}")

    await log_audit(db, user, "bulk_import", "photo", "", f"Photo ZIP upload — {file.filename}",
                    bulk_count=len(matched) + len(matched_staff),
                    metadata={"matched_students": len(matched), "unmatched_students": len(unmatched),
                              "matched_staff": len(matched_staff), "unmatched_staff": len(unmatched_staff)})

    return {
        "matched": len(matched),
        "unmatched": len(unmatched),
        "matched_staff": len(matched_staff),
        "unmatched_staff": len(unmatched_staff),
        "unmatched_names": unmatched[:100],
        "unmatched_staff_names": unmatched_staff[:50],
    }

@router.post("/students/import-student-details")
async def import_student_details(file: UploadFile = File(...), user=Depends(get_current_user), db=Depends(get_tenant_db)):
    """Import student demographic details: teacher, gender, EAL, aboriginal status, NCCD."""
    if user.get("role") not in ["admin", "leadership"]:
        raise HTTPException(403, "Admin or leadership access required")

    content = await file.read()
    fname = (file.filename or "").lower()

    if fname.endswith(".csv"):
        text = content.decode("utf-8-sig")
        all_rows = list(csv.reader(io.StringIO(text)))
    elif fname.endswith(".xlsx") or fname.endswith(".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        all_rows = [[str(cell.value or "").strip() for cell in row] for row in ws.iter_rows()]
    else:
        raise HTTPException(400, "Unsupported file format. Please upload a CSV or XLSX file.")

    if not all_rows:
        raise HTTPException(400, "File is empty.")

    header_idx, headers = 0, []
    for i, row in enumerate(all_rows[:10]):
        cells = [str(v or "").strip().upper() for v in row]
        if "STUDENT_KEY" in cells:
            header_idx, headers = i, cells
            break
    if not headers:
        raise HTTPException(400, "Could not find STUDENT_KEY column. Check the file format.")

    def _col(*names):
        for n in names:
            for i, h in enumerate(headers):
                if h == n.upper() or h.replace(" ", "_") == n.upper().replace(" ", "_"):
                    return i
        return None

    key_c   = _col("STUDENT_KEY")
    home_c  = _col("HOME_GROUP", "Home Group (teacher)")
    gender_c = _col("GENDER", "Gender")
    eal_c   = _col("EAL_STATUS", "EAL Status")
    atsi_c  = _col("ATSI_STATUS", "Aboriginal Status")
    nccd_c  = _col("NCCD_DISABILITY")

    if key_c is None:
        raise HTTPException(400, "STUDENT_KEY column not found in header row.")

    updated, unmatched, errors = 0, [], []

    for row_num, row in enumerate(all_rows[header_idx + 1:], start=header_idx + 2):
        raw = [str(v or "").strip() for v in row]
        if key_c >= len(raw) or not raw[key_c]:
            continue
        student_key = raw[key_c]

        update = {}
        if home_c is not None and home_c < len(raw) and raw[home_c]:
            hg = raw[home_c]
            m = re.match(r'^(\S+)\s*\(([^)]+)\)\s*$', hg)
            if m:
                update["class_name"] = m.group(1)
                update["teacher"] = m.group(2).title()
            else:
                update["class_name"] = hg

        if gender_c is not None and gender_c < len(raw) and raw[gender_c]:
            update["gender"] = raw[gender_c]
        if eal_c is not None and eal_c < len(raw) and raw[eal_c]:
            update["eal_status"] = raw[eal_c]
        if atsi_c is not None and atsi_c < len(raw) and raw[atsi_c]:
            update["aboriginal_status"] = raw[atsi_c]
        if nccd_c is not None and nccd_c < len(raw) and raw[nccd_c]:
            update["nccd_disability"] = raw[nccd_c]

        if not update:
            continue

        result = await db.students.update_one(
            {"$or": [{"sussi_id": student_key}, {"external_id": student_key}]},
            {"$set": update}
        )
        if result.matched_count > 0:
            updated += 1
        else:
            unmatched.append(student_key)

    await log_audit(db, user, "bulk_import", "student", "", f"Student details import — {file.filename}",
                    bulk_count=updated, metadata={"unmatched": len(unmatched)})

    return {
        "updated": updated,
        "unmatched": len(unmatched),
        "unmatched_keys": unmatched[:20],
        "errors": errors,
    }
