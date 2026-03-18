from fastapi import FastAPI, APIRouter, HTTPException, Depends, Response, Request, UploadFile, File
from fastapi.responses import StreamingResponse, RedirectResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.sessions import SessionMiddleware
from authlib.integrations.starlette_client import OAuth
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import uuid
import io
import csv
import json
import random
import httpx
import re
import openpyxl
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Any
from datetime import datetime, timezone, timedelta, date as date_type
from collections import defaultdict

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")
logger = logging.getLogger(__name__)

# Google OAuth
oauth = OAuth()
oauth.register(
    name='google',
    client_id=os.environ['GOOGLE_CLIENT_ID'],
    client_secret=os.environ['GOOGLE_CLIENT_SECRET'],
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    client_kwargs={'scope': 'openid email profile'},
)

# ==============================
# MODELS
# ==============================

class SessionRequest(BaseModel):
    session_id: str

class Student(BaseModel):
    student_id: str = Field(default_factory=lambda: f"stu_{uuid.uuid4().hex[:8]}")
    first_name: str
    preferred_name: Optional[str] = None
    last_name: str
    year_level: str
    class_name: str
    teacher: str = ""
    date_of_birth: Optional[str] = None
    gender: str = ""
    enrolment_status: str = "active"
    external_id: Optional[str] = None  # School system student code for attendance matching
    sussi_id: Optional[str] = None  # SUSSI ID from school system (used for attendance matching)

class AttendanceRecord(BaseModel):
    attendance_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    student_id: str
    date: str
    attendance_status: str
    late_arrival: bool = False
    early_departure: bool = False

class ScreeningSession(BaseModel):
    screening_id: str = Field(default_factory=lambda: f"scr_{uuid.uuid4().hex[:8]}")
    screening_period: str
    year: int = 2025
    date: str
    teacher_id: str
    class_name: str
    status: str = "active"

class SAEBRSResult(BaseModel):
    result_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    student_id: str
    screening_id: str
    social_items: List[int] = []
    academic_items: List[int] = []
    emotional_items: List[int] = []
    social_score: int = 0
    academic_score: int = 0
    emotional_score: int = 0
    total_score: int = 0
    risk_level: str = "Low Risk"
    social_risk: str = "Low Risk"
    academic_risk: str = "Low Risk"
    emotional_risk: str = "Low Risk"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class SAEBRSPlusResult(BaseModel):
    result_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    student_id: str
    screening_id: str
    self_report_items: List[int] = []
    attendance_pct: float = 100.0  # kept for backward compat, no longer collected at screening
    social_domain: int = 0
    academic_domain: int = 0
    emotional_domain: int = 0
    belonging_domain: int = 0
    wellbeing_total: int = 0
    wellbeing_tier: int = 1
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Intervention(BaseModel):
    intervention_id: str = Field(default_factory=lambda: f"int_{uuid.uuid4().hex[:8]}")
    student_id: str
    intervention_type: str
    assigned_staff: str
    start_date: str
    review_date: str
    status: str = "active"
    goals: str = ""
    progress_notes: str = ""
    frequency: str = ""
    outcome_rating: Optional[int] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class CaseNote(BaseModel):
    case_id: str = Field(default_factory=lambda: f"case_{uuid.uuid4().hex[:8]}")
    student_id: str
    staff_member: str
    date: str
    note_type: str
    notes: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class SchoolSettings(BaseModel):
    school_name: str = "Demo School"
    school_type: str = "both"
    current_term: str = "Term 1"
    current_year: int = 2025

# ==============================
# HELPER FUNCTIONS
# ==============================

def compute_saebrs_risk(total: int, social: int, academic: int, emotional: int, thresholds: dict = None):
    t = thresholds or {}
    t_some = t.get("saebrs_some_risk", 37)
    t_high = t.get("saebrs_high_risk", 24)
    total_risk = "Low Risk" if total >= t_some else ("Some Risk" if total >= t_high else "High Risk")
    social_risk = "Low Risk" if social >= 13 else ("Some Risk" if social >= 8 else "High Risk")
    academic_risk = "Low Risk" if academic >= 10 else ("Some Risk" if academic >= 6 else "High Risk")
    emotional_risk = "Low Risk" if emotional >= 16 else ("Some Risk" if emotional >= 12 else "High Risk")
    return total_risk, social_risk, academic_risk, emotional_risk

def compute_wellbeing_tier(total: int) -> int:
    if total >= 50: return 1
    elif total >= 35: return 2
    return 3

def compute_attendance_score(pct: float) -> int:
    if pct >= 95: return 3
    elif pct >= 90: return 2
    elif pct >= 80: return 1
    return 0

def compute_mtss_tier(saebrs_risk: str, wellbeing_tier: int, attendance_pct: float, thresholds: dict = None) -> int:
    t = thresholds or {}
    att_high = t.get("attendance_high_risk", 80.0)
    att_some = t.get("attendance_some_risk", 90.0)
    if saebrs_risk == "High Risk" or wellbeing_tier == 3 or attendance_pct < att_high:
        return 3
    elif saebrs_risk == "Some Risk" or wellbeing_tier == 2 or attendance_pct < att_some:
        return 2
    return 1

async def get_school_settings_doc():
    s = await db.school_settings.find_one({}, {"_id": 0})
    return s or {}

SETTINGS_DEFAULTS = {
    "school_name": "", "school_type": "both", "current_term": "Term 1", "current_year": 2025,
    "platform_name": "WellTrack", "logo_base64": "", "accent_color": "#0f172a", "welcome_message": "",
    "tier_thresholds": {"saebrs_some_risk": 37, "saebrs_high_risk": 24, "attendance_some_risk": 90.0, "attendance_high_risk": 80.0},
    "modules_enabled": {"saebrs_plus": True},
    "intervention_types": ["Counselling", "Behaviour Support", "Social Skills Groups", "Mentoring", "Academic Support", "Attendance Intervention", "Check-In/Check-Out", "Parent Consultation", "Peer Mentoring", "Referral – External Services"],
    "year_start_month": 2,
    "custom_student_fields": [],
    "risk_config": {"consecutive_absence_days": 3},
    "ollama_url": "http://localhost:11434",
    "ollama_model": "llama3.2",
    "ai_suggestions_enabled": True,
    "excluded_absence_types": [],
}

async def get_current_user(request: Request):
    session_token = request.cookies.get("session_token")
    if not session_token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            session_token = auth_header.split(" ")[1]
    if not session_token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    session_doc = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
    if not session_doc:
        raise HTTPException(status_code=401, detail="Invalid session")
    expires_at = session_doc["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expired")
    user_doc = await db.users.find_one({"user_id": session_doc["user_id"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="User not found")
    return user_doc

async def create_alert(student_id: str, student_name: str, class_name: str, alert_type: str, severity: str, message: str):
    existing = await db.alerts.find_one({"student_id": student_id, "alert_type": alert_type, "resolved": False})
    if existing:
        return
    await db.alerts.insert_one({
        "alert_id": f"alrt_{uuid.uuid4().hex[:8]}",
        "student_id": student_id, "student_name": student_name, "class_name": class_name,
        "alert_type": alert_type, "severity": severity, "message": message,
        "created_at": datetime.now(timezone.utc).isoformat(), "is_read": False, "resolved": False
    })

PRESENT_STATUSES = {
    "Present", "Late arrival, approved", "Late arrival at School",
    "Early departure, approved", "Early departure from School",
}

FULL_PRESENT_STATUSES = {
    "Late arrival, approved", "Late arrival at School",
    "Early departure, approved", "Early departure from School",
}

async def get_student_attendance_pct(student_id: str) -> float:
    stats = await get_student_attendance_stats(student_id)
    return stats["pct"]

async def get_student_attendance_stats(student_id: str) -> dict:
    """Returns {pct, total_sessions, absent_sessions} for a student."""
    school_days_list = await db.school_days.distinct("date")

    if school_days_list:
        exc_records = await db.attendance_records.find({"student_id": student_id}, {"_id": 0}).to_list(5000)
        exc_by_date = {r["date"]: r for r in exc_records}
        settings_doc = await get_school_settings_doc()
        excluded_types = set(settings_doc.get("excluded_absence_types", []))

        total_sessions = 0
        present_sessions = 0.0
        absent_sessions = 0

        for day in school_days_list:
            if day not in exc_by_date:
                total_sessions += 2
                present_sessions += 2.0
            else:
                rec = exc_by_date[day]
                for sess in [rec.get("am_status", ""), rec.get("pm_status", "")]:
                    s = (sess or "").strip()
                    if not s:
                        total_sessions += 1
                        present_sessions += 1.0
                    elif s in excluded_types:
                        pass  # excluded — neutral
                    elif s == "Present":
                        total_sessions += 1
                        present_sessions += 0.5
                        absent_sessions += 1
                    elif s in FULL_PRESENT_STATUSES:
                        total_sessions += 1
                        present_sessions += 1.0
                    else:
                        total_sessions += 1
                        present_sessions += 0.0
                        absent_sessions += 1

        if total_sessions == 0:
            return {"pct": 100.0, "total_sessions": 0, "absent_sessions": 0}
        pct = (present_sessions / total_sessions) * 100.0
        return {"pct": pct, "total_sessions": total_sessions, "absent_sessions": absent_sessions}
    else:
        # Fallback: calculate from raw attendance_records
        records = await db.attendance_records.find({"student_id": student_id}, {"_id": 0}).to_list(5000)
        if not records:
            return {"pct": 100.0, "total_sessions": 0, "absent_sessions": 0}
        total_sessions = 0
        absent_sessions = 0
        for r in records:
            am = (r.get("am_status") or "").strip()
            pm = (r.get("pm_status") or "").strip()
            if am:
                total_sessions += 1
                if am not in PRESENT_STATUSES:
                    absent_sessions += 1
            if pm:
                total_sessions += 1
                if pm not in PRESENT_STATUSES:
                    absent_sessions += 1
        if total_sessions == 0:
            return {"pct": 100.0, "total_sessions": 0, "absent_sessions": 0}
        pct = ((total_sessions - absent_sessions) / total_sessions) * 100
        return {"pct": pct, "total_sessions": total_sessions, "absent_sessions": absent_sessions}

# ==============================
# AUTH ROUTES
# ==============================

# ==============================
# AUTH ROUTES
# ==============================

@api_router.get("/auth/google")
async def google_login(request: Request):
    from urllib.parse import urlencode
    state = uuid.uuid4().hex
    # Store state in MongoDB — avoids Starlette session-cookie issues in proxied/K8s environments
    await db.oauth_states.insert_one({
        "state": state,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=10)).isoformat(),
    })
    params = urlencode({
        "response_type": "code",
        "client_id": os.environ['GOOGLE_CLIENT_ID'],
        "redirect_uri": os.environ['GOOGLE_REDIRECT_URI'],
        "scope": "openid email profile",
        "state": state,
        "prompt": "select_account",
    })
    return RedirectResponse(url=f"https://accounts.google.com/o/oauth2/v2/auth?{params}")

@api_router.get("/auth/callback")
async def google_callback(request: Request):
    frontend_url = os.environ['FRONTEND_URL']
    code = request.query_params.get("code")
    state = request.query_params.get("state")
    error = request.query_params.get("error")

    if error:
        logger.error(f"Google OAuth error: {error}")
        return RedirectResponse(url=f"{frontend_url}/login?error=access_denied")

    if not code or not state:
        return RedirectResponse(url=f"{frontend_url}/login?error=auth_failed")

    # Verify + consume state from DB (no session cookies needed)
    state_doc = await db.oauth_states.find_one_and_delete({"state": state})
    if not state_doc:
        logger.error(f"OAuth state not found or already used: {state}")
        return RedirectResponse(url=f"{frontend_url}/login?error=auth_failed")

    expires_at_str = state_doc.get("expires_at", "")
    if expires_at_str:
        exp = datetime.fromisoformat(expires_at_str).replace(tzinfo=timezone.utc)
        if exp < datetime.now(timezone.utc):
            return RedirectResponse(url=f"{frontend_url}/login?error=auth_failed")

    # Exchange code for tokens via Google's token endpoint
    try:
        async with httpx.AsyncClient() as hc:
            tok_resp = await hc.post(
                "https://oauth2.googleapis.com/token",
                data={
                    "code": code,
                    "client_id": os.environ['GOOGLE_CLIENT_ID'],
                    "client_secret": os.environ['GOOGLE_CLIENT_SECRET'],
                    "redirect_uri": os.environ['GOOGLE_REDIRECT_URI'],
                    "grant_type": "authorization_code",
                },
            )
        if tok_resp.status_code != 200:
            logger.error(f"Token exchange failed: {tok_resp.text}")
            return RedirectResponse(url=f"{frontend_url}/login?error=auth_failed")

        access_token = tok_resp.json().get("access_token")

        async with httpx.AsyncClient() as hc:
            ui_resp = await hc.get(
                "https://www.googleapis.com/oauth2/v3/userinfo",
                headers={"Authorization": f"Bearer {access_token}"},
            )
        if ui_resp.status_code != 200:
            logger.error(f"Userinfo request failed: {ui_resp.text}")
            return RedirectResponse(url=f"{frontend_url}/login?error=auth_failed")

        user_info = ui_resp.json()
    except Exception as e:
        logger.error(f"OAuth flow error: {e}")
        return RedirectResponse(url=f"{frontend_url}/login?error=auth_failed")

    email = user_info.get('email', '').lower().strip()
    name = user_info.get('name', '')
    picture = user_info.get('picture', '')

    if not email:
        return RedirectResponse(url=f"{frontend_url}/login?error=no_email")

    # Look up user first, then decide whether to create or deny
    existing = await db.users.find_one({"email": email}, {"_id": 0})
    user_count = await db.users.count_documents({})

    if user_count == 0:
        # First-ever user — auto-register as school administrator
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        new_user = {
            "user_id": user_id, "email": email, "name": name,
            "picture": picture, "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.users.insert_one({**new_user})
        existing = new_user
    elif not existing:
        return RedirectResponse(url=f"{frontend_url}/login?error=access_denied")
    else:
        await db.users.update_one({"email": email}, {"$set": {"name": name, "picture": picture}})
        existing = {**existing, "name": name, "picture": picture}

    user_id = existing["user_id"]
    session_token = f"sess_{uuid.uuid4().hex}"
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    await db.user_sessions.insert_one({
        "user_id": user_id, "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat(),
    })

    # Redirect to onboarding if not yet complete, otherwise dashboard
    settings_doc = await db.school_settings.find_one({}, {"_id": 0})
    onboarding_complete = bool(settings_doc and settings_doc.get("onboarding_complete"))
    redirect_target = "dashboard" if onboarding_complete else "onboarding"

    redirect = RedirectResponse(url=f"{frontend_url}/{redirect_target}")
    redirect.set_cookie(
        key="session_token", value=session_token,
        httponly=True, secure=True, samesite="none", path="/", max_age=7*24*3600,
    )
    return redirect

@api_router.get("/auth/me")
async def get_me(user=Depends(get_current_user)):
    return user

# ==============================
# ONBOARDING
# ==============================

@api_router.get("/onboarding/status")
async def get_onboarding_status():
    """Public endpoint — no auth required."""
    settings = await db.school_settings.find_one({}, {"_id": 0})
    user_count = await db.users.count_documents({})
    return {
        "complete": bool(settings and settings.get("onboarding_complete", False)),
        "has_users": user_count > 0,
        "school_name": settings.get("school_name", "") if settings else "",
    }

@api_router.post("/onboarding/complete")
async def complete_onboarding(data: dict, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    await db.school_settings.update_one(
        {},
        {"$set": {
            "school_name": data.get("school_name", "My School"),
            "school_type": data.get("school_type", "both"),
            "current_term": data.get("current_term", "Term 1"),
            "current_year": data.get("current_year", 2025),
            "onboarding_complete": True,
        }},
        upsert=True
    )
    return {"message": "Onboarding complete"}

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    token = request.cookies.get("session_token")
    if token:
        await db.user_sessions.delete_one({"session_token": token})
    response.delete_cookie("session_token", path="/", samesite="none", secure=True)
    return {"message": "Logged out"}

@api_router.put("/auth/role")
async def update_role(data: dict, user=Depends(get_current_user)):
    # Only admins can change roles
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only administrators can change user roles")
    target_user_id = data.get("user_id", user["user_id"])
    role = data.get("role")
    if role not in ["teacher", "wellbeing", "leadership", "admin"]:
        raise HTTPException(status_code=400, detail="Invalid role")
    await db.users.update_one({"user_id": target_user_id}, {"$set": {"role": role}})
    return {"message": "Role updated", "role": role}

# ==============================
# USER MANAGEMENT (Admin only)
# ==============================

@api_router.get("/users")
async def get_users(user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    users = await db.users.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    return users

@api_router.post("/users")
async def create_user(data: dict, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    email = data.get("email", "").lower().strip()
    name = data.get("name", "")
    role = data.get("role", "teacher")
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=409, detail="User with this email already exists")
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    new_user = {
        "user_id": user_id, "email": email, "name": name,
        "picture": "", "role": role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one({**new_user})
    return new_user

@api_router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, data: dict, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    role = data.get("role")
    if role not in ["teacher", "wellbeing", "leadership", "admin", "screener"]:
        raise HTTPException(status_code=400, detail="Invalid role")
    await db.users.update_one({"user_id": user_id}, {"$set": {"role": role}})
    return {"message": "Role updated"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    if user_id == user["user_id"]:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    await db.users.delete_one({"user_id": user_id})
    await db.user_sessions.delete_many({"user_id": user_id})
    return {"message": "User deleted"}

# ==============================
# STUDENT ROUTES
# ==============================

@api_router.get("/students")
async def get_students(class_name: Optional[str] = None, year_level: Optional[str] = None, user=Depends(get_current_user)):
    query = {"enrolment_status": "active"}
    if class_name:
        query["class_name"] = class_name
    if year_level:
        query["year_level"] = year_level
    students = await db.students.find(query, {"_id": 0}).sort("last_name", 1).to_list(500)
    return students

@api_router.post("/students")
async def create_student(student: Student, user=Depends(get_current_user)):
    d = student.model_dump()
    await db.students.insert_one({**d})
    return d

@api_router.post("/students/import")
async def import_students(data: dict, user=Depends(get_current_user)):
    """Import students from new school system export format (SussiId-based CSV).
    Upserts by sussi_id to avoid duplicates. Also handles legacy format.
    """
    rows = data.get("students", [])
    if not rows:
        raise HTTPException(status_code=400, detail="No student data provided")

    imported, updated, errors = [], [], []
    for i, row in enumerate(rows):
        # Skip non-student or inactive rows
        base_role = str(row.get("Base Role") or row.get("base_role") or "Student").strip()
        user_status = str(row.get("User Status") or row.get("user_status") or "Active").strip()
        if base_role.lower() not in ("student", "") or user_status.lower() == "inactive":
            continue

        # Map new format columns
        sussi_id = str(row.get("SussiId") or row.get("sussi_id") or row.get("Import Identifier") or "").strip()
        first_name = str(row.get("First Name") or row.get("first_name") or "").strip()
        preferred_name = str(row.get("Preferred Name") or row.get("preferred_name") or "").strip()
        last_name = str(row.get("Surname") or row.get("last_name") or "").strip()
        class_name = str(row.get("Form Group") or row.get("class_name") or "").strip()
        year_level = str(row.get("Year Level") or row.get("year_level") or "").strip()
        teacher = str(row.get("teacher") or "").strip()
        gender = str(row.get("gender") or "").strip()
        dob = str(row.get("date_of_birth") or "").strip()

        if not sussi_id and not first_name and not last_name:
            errors.append({"row": i + 1, "error": "Missing student identifier (SussiId, First Name, or Surname)"})
            continue

        student_doc = {
            "first_name": first_name or sussi_id,
            "preferred_name": preferred_name or None,
            "last_name": last_name,
            "year_level": year_level,
            "class_name": class_name,
            "teacher": teacher,
            "gender": gender,
            "date_of_birth": dob,
            "enrolment_status": "active",
            "sussi_id": sussi_id,
            "external_id": sussi_id,  # Use SussiId for attendance matching
        }

        # Upsert by sussi_id if available, otherwise insert
        if sussi_id:
            existing = await db.students.find_one({"sussi_id": sussi_id})
            if existing:
                await db.students.update_one({"sussi_id": sussi_id}, {"$set": student_doc})
                updated.append(sussi_id)
            else:
                student_doc["student_id"] = f"stu_{uuid.uuid4().hex[:8]}"
                await db.students.insert_one({**student_doc})
                imported.append(student_doc)
        else:
            student_doc["student_id"] = f"stu_{uuid.uuid4().hex[:8]}"
            await db.students.insert_one({**student_doc})
            imported.append(student_doc)

    return {"imported": len(imported), "updated": len(updated), "errors": errors, "total": len(rows)}

@api_router.get("/students/summary")
async def get_students_summary(class_name: Optional[str] = None, year_level: Optional[str] = None, user=Depends(get_current_user)):
    query = {"enrolment_status": "active"}
    if class_name:
        query["class_name"] = class_name
    if year_level:
        query["year_level"] = year_level
    students = await db.students.find(query, {"_id": 0}).sort("last_name", 1).to_list(500)

    result = []
    for s in students:
        sid = s["student_id"]
        saebrs = await db.saebrs_results.find_one({"student_id": sid}, {"_id": 0}, sort=[("created_at", -1)])
        plus = await db.saebrs_plus_results.find_one({"student_id": sid}, {"_id": 0}, sort=[("created_at", -1)])
        att_pct = await get_student_attendance_pct(sid)
        active_interventions = await db.interventions.count_documents({"student_id": sid, "status": "active"})

        if saebrs and plus:
            tier = compute_mtss_tier(saebrs["risk_level"], plus["wellbeing_tier"], att_pct)
        elif saebrs:
            tier = 3 if saebrs["risk_level"] == "High Risk" else (2 if saebrs["risk_level"] == "Some Risk" else 1)
        else:
            tier = None

        result.append({
            **s,
            "mtss_tier": tier,
            "saebrs_risk": saebrs["risk_level"] if saebrs else "Not Screened",
            "saebrs_total": saebrs["total_score"] if saebrs else None,
            "wellbeing_tier": plus["wellbeing_tier"] if plus else None,
            "wellbeing_total": plus["wellbeing_total"] if plus else None,
            "attendance_pct": round(att_pct, 1),
            "active_interventions": active_interventions
        })
    return result

@api_router.get("/students/{student_id}")
async def get_student(student_id: str, user=Depends(get_current_user)):
    s = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Student not found")
    return s

@api_router.get("/students/{student_id}/profile")
async def get_student_profile(student_id: str, user=Depends(get_current_user)):
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    saebrs_results = await db.saebrs_results.find({"student_id": student_id}, {"_id": 0}).sort("created_at", 1).to_list(20)
    saebrs_plus = await db.saebrs_plus_results.find({"student_id": student_id}, {"_id": 0}).sort("created_at", 1).to_list(20)
    interventions = await db.interventions.find({"student_id": student_id}, {"_id": 0}).sort("created_at", -1).to_list(20)
    case_notes = await db.case_notes.find({"student_id": student_id}, {"_id": 0}).sort("date", -1).to_list(20)
    att_pct = await get_student_attendance_pct(student_id)
    attendance_records = await db.attendance.find({"student_id": student_id}, {"_id": 0}).sort("date", -1).to_list(50)
    alerts = await db.alerts.find({"student_id": student_id, "resolved": False}, {"_id": 0}).to_list(10)

    latest_saebrs = saebrs_results[-1] if saebrs_results else None
    latest_plus = saebrs_plus[-1] if saebrs_plus else None

    if latest_saebrs and latest_plus:
        tier = compute_mtss_tier(latest_saebrs["risk_level"], latest_plus["wellbeing_tier"], att_pct)
    elif latest_saebrs:
        tier = 3 if latest_saebrs["risk_level"] == "High Risk" else (2 if latest_saebrs["risk_level"] == "Some Risk" else 1)
    else:
        tier = None

    return {
        "student": student, "mtss_tier": tier,
        "attendance_pct": round(att_pct, 1),
        "saebrs_results": saebrs_results,
        "saebrs_plus_results": saebrs_plus,
        "interventions": interventions,
        "case_notes": case_notes,
        "attendance_records": attendance_records,
        "alerts": alerts
    }

# ==============================
# ATTENDANCE ROUTES
# ==============================


# ==============================
# SCREENING ROUTES
# ==============================

@api_router.get("/screening/sessions")
async def get_sessions(class_name: Optional[str] = None, user=Depends(get_current_user)):
    query = {}
    if class_name:
        query["class_name"] = {"$in": [class_name, "all"]}
    sessions = await db.screening_sessions.find(query, {"_id": 0}).sort("date", -1).to_list(100)
    return sessions

@api_router.post("/screening/sessions")
async def create_session(session: ScreeningSession, user=Depends(get_current_user)):
    d = session.model_dump()
    await db.screening_sessions.insert_one({**d})
    return d

@api_router.post("/screening/saebrs")
async def submit_saebrs(result: SAEBRSResult, user=Depends(get_current_user)):
    s = sum(result.social_items) if result.social_items else result.social_score
    a = sum(result.academic_items) if result.academic_items else result.academic_score
    e = sum(result.emotional_items) if result.emotional_items else result.emotional_score
    t = s + a + e
    school_s = await get_school_settings_doc()
    thresholds = school_s.get("tier_thresholds", {})
    r, sr, ar, er = compute_saebrs_risk(t, s, a, e, thresholds)
    d = result.model_dump()
    d.update({"social_score": s, "academic_score": a, "emotional_score": e,
               "total_score": t, "risk_level": r, "social_risk": sr, "academic_risk": ar, "emotional_risk": er})

    # Tier change alert: compare with previous screening
    prev_saebrs = await db.saebrs_results.find_one({"student_id": result.student_id}, {"_id": 0}, sort=[("created_at", -1)])
    prev_plus = await db.saebrs_plus_results.find_one({"student_id": result.student_id}, {"_id": 0}, sort=[("created_at", -1)])
    att_pct = await get_student_attendance_pct(result.student_id)

    if prev_saebrs:
        old_tier = compute_mtss_tier(prev_saebrs["risk_level"], prev_plus["wellbeing_tier"] if prev_plus else 1, att_pct, thresholds)
        new_tier = compute_mtss_tier(r, prev_plus["wellbeing_tier"] if prev_plus else 1, att_pct, thresholds)
        if old_tier != new_tier:
            student = await db.students.find_one({"student_id": result.student_id}, {"_id": 0})
            if student:
                name = f"{student['first_name']} {student['last_name']}"
                severity = "high" if new_tier == 3 else "medium"
                await db.alerts.insert_one({
                    "alert_id": f"alrt_{uuid.uuid4().hex[:8]}",
                    "student_id": result.student_id, "student_name": name, "class_name": student["class_name"],
                    "alert_type": "tier_change",
                    "severity": severity,
                    "message": f"{name} moved from Tier {old_tier} to Tier {new_tier} — pending approval",
                    "previous_tier": old_tier, "new_tier": new_tier,
                    "pending_approval": True,
                    "created_at": datetime.now(timezone.utc).isoformat(), "is_read": False, "resolved": False
                })

    await db.saebrs_results.insert_one({**d})
    student = await db.students.find_one({"student_id": result.student_id}, {"_id": 0})
    if student and r == "High Risk":
        name = f"{student['first_name']} {student['last_name']}"
        await create_alert(result.student_id, name, student["class_name"], "high_risk_saebrs", "high",
                           f"{name} screened as High Risk (SAEBRS total: {t}/57)")
    return d

@api_router.post("/screening/saebrs-plus")
async def submit_saebrs_plus(result: SAEBRSPlusResult, user=Depends(get_current_user)):
    saebrs = await db.saebrs_results.find_one({"student_id": result.student_id, "screening_id": result.screening_id}, {"_id": 0}, sort=[("created_at", -1)])
    soc = saebrs["social_score"] if saebrs else result.social_domain
    aca = saebrs["academic_score"] if saebrs else result.academic_domain
    items = result.self_report_items
    if len(items) >= 7:
        item0_rev = 3 - items[0]
        emo = item0_rev + items[1] + items[2]
        bel = items[3] + items[4] + items[5] + items[6]
        if items[0] >= 2:
            student = await db.students.find_one({"student_id": result.student_id}, {"_id": 0})
            if student:
                name = f"{student['first_name']} {student['last_name']}"
                await create_alert(result.student_id, name, student["class_name"], "emotional_distress", "high",
                                   f"{name} reported high emotional distress in self-assessment")
    else:
        emo = result.emotional_domain
        bel = result.belonging_domain
    # Attendance no longer included in wellbeing total — tracked separately
    total = soc + aca + emo + bel
    tier = compute_wellbeing_tier(total)
    d = result.model_dump()
    d.update({"social_domain": soc, "academic_domain": aca, "emotional_domain": emo,
               "belonging_domain": bel, "wellbeing_total": total, "wellbeing_tier": tier})
    await db.saebrs_plus_results.insert_one({**d})
    return d

@api_router.get("/screening/results/{student_id}")
async def get_screening_results(student_id: str, user=Depends(get_current_user)):
    saebrs = await db.saebrs_results.find({"student_id": student_id}, {"_id": 0}).sort("created_at", 1).to_list(20)
    plus = await db.saebrs_plus_results.find({"student_id": student_id}, {"_id": 0}).sort("created_at", 1).to_list(20)
    return {"saebrs": saebrs, "saebrs_plus": plus}

# ==============================
# INTERVENTION ROUTES
# ==============================

@api_router.get("/interventions")
async def get_interventions(student_id: Optional[str] = None, status: Optional[str] = None, user=Depends(get_current_user)):
    query = {}
    if student_id:
        query["student_id"] = student_id
    if status:
        query["status"] = status
    items = await db.interventions.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items

@api_router.post("/interventions")
async def create_intervention(intervention: Intervention, user=Depends(get_current_user)):
    d = intervention.model_dump()
    await db.interventions.insert_one({**d})
    return d

@api_router.put("/interventions/{iid}")
async def update_intervention(iid: str, data: dict, user=Depends(get_current_user)):
    await db.interventions.update_one({"intervention_id": iid}, {"$set": data})
    return await db.interventions.find_one({"intervention_id": iid}, {"_id": 0})

@api_router.delete("/interventions/{iid}")
async def delete_intervention(iid: str, user=Depends(get_current_user)):
    await db.interventions.delete_one({"intervention_id": iid})
    return {"message": "Deleted"}

@api_router.post("/interventions/ai-suggest/{student_id}")
async def get_ai_suggestions(student_id: str, user=Depends(get_current_user)):
    """Get AI-powered intervention suggestions via Ollama local LLM."""
    settings_doc = await get_school_settings_doc()
    if not settings_doc.get("ai_suggestions_enabled", True):
        raise HTTPException(403, "AI suggestions are disabled. Enable in Settings > Integrations.")

    ollama_url = settings_doc.get("ollama_url", "http://localhost:11434")
    ollama_model = settings_doc.get("ollama_model", "llama3.2")

    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(404, "Student not found")

    latest_saebrs = await db.saebrs_results.find_one({"student_id": student_id}, {"_id": 0}, sort=[("created_at", -1)])
    att_pct = await get_student_attendance_pct(student_id)
    active_ints = await db.interventions.find({"student_id": student_id, "status": "active"}, {"_id": 0}).to_list(10)

    display_name = student.get('first_name', '')
    if student.get('preferred_name'):
        display_name += f" ({student['preferred_name']})"
    display_name += f" {student.get('last_name', '')}"

    context = f"Student: {display_name.strip()}, Year Level: {student.get('year_level', 'Unknown')}\n"
    if latest_saebrs:
        context += f"SAEBRS: {latest_saebrs['total_score']}/57 ({latest_saebrs['risk_level']}) - Social: {latest_saebrs.get('social_score', 0)}, Academic: {latest_saebrs.get('academic_score', 0)}, Emotional: {latest_saebrs.get('emotional_score', 0)}\n"
    else:
        context += "SAEBRS: Not screened\n"
    context += f"Attendance: {att_pct:.1f}%\n"
    context += f"Current interventions: {[i['intervention_type'] for i in active_ints] or ['None']}\n"

    prompt = f"""You are an MTSS (Multi-Tiered System of Supports) specialist. Based on this student's data, suggest 3 evidence-based interventions.

{context}

Respond with ONLY a valid JSON array of exactly 3 objects, each with these keys:
- "type": intervention name (string)
- "priority": "high", "medium", or "low"
- "rationale": brief explanation (1-2 sentences)
- "goals": measurable goal (1 sentence)
- "frequency": how often (e.g. "Weekly", "3x per week")
- "timeline": duration (e.g. "8 weeks", "1 term")

Example format: [{{"type": "Counselling", "priority": "high", "rationale": "...", "goals": "...", "frequency": "Weekly", "timeline": "8 weeks"}}]"""

    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(
                f"{ollama_url}/api/generate",
                json={"model": ollama_model, "prompt": prompt, "stream": False}
            )
            resp.raise_for_status()
            result = resp.json()
            content = result.get("response", "")
            json_match = re.search(r'\[.*\]', content, re.DOTALL)
            if json_match:
                recs = json.loads(json_match.group())
                return {"recommendations": recs[:3]}
            return {"recommendations": []}
    except httpx.ConnectError:
        raise HTTPException(503, f"Cannot connect to Ollama at {ollama_url}. Ensure Ollama is running and accessible.")
    except Exception as e:
        raise HTTPException(500, f"AI service error: {str(e)}")


# ==============================
# CASE NOTES
# ==============================

@api_router.get("/case-notes")
async def get_case_notes(student_id: Optional[str] = None, user=Depends(get_current_user)):
    q = {}
    if student_id:
        q["student_id"] = student_id
    return await db.case_notes.find(q, {"_id": 0}).sort("date", -1).to_list(500)

@api_router.post("/case-notes")
async def add_case_note(note: CaseNote, user=Depends(get_current_user)):
    d = note.model_dump()
    await db.case_notes.insert_one({**d})
    return d

@api_router.delete("/case-notes/{case_id}")
async def delete_case_note(case_id: str, user=Depends(get_current_user)):
    await db.case_notes.delete_one({"case_id": case_id})
    return {"message": "Deleted"}

@api_router.put("/case-notes/{case_id}")
async def update_case_note(case_id: str, data: dict, user=Depends(get_current_user)):
    data.pop("_id", None)
    data.pop("case_id", None)
    await db.case_notes.update_one({"case_id": case_id}, {"$set": data})
    return await db.case_notes.find_one({"case_id": case_id}, {"_id": 0})

# ==============================
# ALERTS
# ==============================

@api_router.get("/alerts")
async def get_alerts(resolved: bool = False, user=Depends(get_current_user)):
    return await db.alerts.find({"resolved": resolved}, {"_id": 0}).sort("created_at", -1).to_list(200)

@api_router.put("/alerts/{alert_id}/read")
async def mark_read(alert_id: str, user=Depends(get_current_user)):
    await db.alerts.update_one({"alert_id": alert_id}, {"$set": {"is_read": True}})
    return {"message": "ok"}

@api_router.put("/alerts/{alert_id}/resolve")
async def resolve_alert(alert_id: str, user=Depends(get_current_user)):
    await db.alerts.update_one({"alert_id": alert_id}, {"$set": {"resolved": True, "is_read": True}})
    return {"message": "ok"}

@api_router.put("/alerts/read-all")
async def mark_all_read(user=Depends(get_current_user)):
    await db.alerts.update_many({"is_read": False, "resolved": False}, {"$set": {"is_read": True}})
    return {"message": "ok"}

# ==============================
# ANALYTICS
# ==============================

@api_router.get("/analytics/tier-distribution")
async def tier_distribution(user=Depends(get_current_user)):
    students = await db.students.find({"enrolment_status": "active"}, {"_id": 0}).to_list(500)
    counts = {"tier1": 0, "tier2": 0, "tier3": 0, "unscreened": 0}
    class_breakdown = {}
    for s in students:
        sid = s["student_id"]
        saebrs = await db.saebrs_results.find_one({"student_id": sid}, {"_id": 0}, sort=[("created_at", -1)])
        plus = await db.saebrs_plus_results.find_one({"student_id": sid}, {"_id": 0}, sort=[("created_at", -1)])
        att_pct = await get_student_attendance_pct(sid)
        cls = s["class_name"]
        if cls not in class_breakdown:
            class_breakdown[cls] = {"tier1": 0, "tier2": 0, "tier3": 0, "teacher": s["teacher"]}
        if saebrs and plus:
            t = compute_mtss_tier(saebrs["risk_level"], plus["wellbeing_tier"], att_pct)
            counts[f"tier{t}"] += 1
            class_breakdown[cls][f"tier{t}"] += 1
        elif saebrs:
            t = 3 if saebrs["risk_level"] == "High Risk" else (2 if saebrs["risk_level"] == "Some Risk" else 1)
            counts[f"tier{t}"] += 1
            class_breakdown[cls][f"tier{t}"] += 1
        else:
            counts["unscreened"] += 1
    return {"tier_distribution": counts, "total_students": len(students), "class_breakdown": class_breakdown}

@api_router.get("/analytics/classroom-radar/{class_name}")
async def classroom_radar(class_name: str, user=Depends(get_current_user)):
    students = await db.students.find({"class_name": class_name, "enrolment_status": "active"}, {"_id": 0}).to_list(50)
    result = []
    for s in students:
        sid = s["student_id"]
        all_saebrs = await db.saebrs_results.find({"student_id": sid}, {"_id": 0}).sort("created_at", 1).to_list(10)
        saebrs = all_saebrs[-1] if all_saebrs else None
        plus = await db.saebrs_plus_results.find_one({"student_id": sid}, {"_id": 0}, sort=[("created_at", -1)])
        att_pct = await get_student_attendance_pct(sid)
        score_trend = None
        if len(all_saebrs) >= 2:
            score_trend = all_saebrs[-1]["total_score"] - all_saebrs[-2]["total_score"]
        if saebrs and plus:
            tier = compute_mtss_tier(saebrs["risk_level"], plus["wellbeing_tier"], att_pct)
        elif saebrs:
            tier = 3 if saebrs["risk_level"] == "High Risk" else (2 if saebrs["risk_level"] == "Some Risk" else 1)
        else:
            tier = 0
        indicators = []
        if plus and plus.get("belonging_domain", 12) <= 4: indicators.append("low_belonging")
        if plus and plus.get("emotional_domain", 9) <= 3: indicators.append("emotional_distress")
        if att_pct < 90: indicators.append("attendance_decline")
        if score_trend is not None and score_trend <= -8: indicators.append("rapid_score_drop")
        if saebrs and saebrs.get("social_risk") != "Low Risk": indicators.append("social_behaviour_risk")
        if saebrs and saebrs.get("academic_risk") != "Low Risk": indicators.append("academic_engagement_risk")
        active_int = await db.interventions.find_one({"student_id": sid, "status": "active"}, {"_id": 0})
        result.append({
            "student": s, "mtss_tier": tier,
            "saebrs_risk": saebrs["risk_level"] if saebrs else "Not Screened",
            "saebrs_total": saebrs["total_score"] if saebrs else None,
            "social_risk": saebrs["social_risk"] if saebrs else None,
            "academic_risk": saebrs["academic_risk"] if saebrs else None,
            "emotional_risk": saebrs["emotional_risk"] if saebrs else None,
            "wellbeing_tier": plus["wellbeing_tier"] if plus else None,
            "wellbeing_total": plus["wellbeing_total"] if plus else None,
            "belonging_domain": plus["belonging_domain"] if plus else None,
            "emotional_domain": plus["emotional_domain"] if plus else None,
            "attendance_pct": round(att_pct, 1), "score_trend": score_trend,
            "risk_indicators": indicators,
            "has_active_intervention": bool(active_int)
        })
    result.sort(key=lambda x: (-(x["mtss_tier"] or 0), -len(x["risk_indicators"])))
    return result

@api_router.get("/analytics/school-wide")
async def school_wide(year_level: Optional[str] = None, user=Depends(get_current_user)):
    query = {"enrolment_status": "active"}
    if year_level:
        query["year_level"] = year_level
    students = await db.students.find(query, {"_id": 0}).to_list(500)
    domain_totals = {"social": 0, "academic": 0, "emotional": 0, "belonging": 0, "attendance": 0}
    domain_counts = 0
    tier_by_year = {}
    tier_distribution = {1: 0, 2: 0, 3: 0}
    risk_distribution = {"low": 0, "some": 0, "high": 0}
    class_breakdown = {}
    screened_count = 0

    for s in students:
        sid = s["student_id"]
        yr = s["year_level"]
        cls = s["class_name"]

        if yr not in tier_by_year:
            tier_by_year[yr] = {"tier1": 0, "tier2": 0, "tier3": 0}
        if cls not in class_breakdown:
            class_breakdown[cls] = {"tier1": 0, "tier2": 0, "tier3": 0}

        saebrs = await db.saebrs_results.find_one({"student_id": sid}, {"_id": 0}, sort=[("created_at", -1)])
        plus = await db.saebrs_plus_results.find_one({"student_id": sid}, {"_id": 0}, sort=[("created_at", -1)])
        att_pct = await get_student_attendance_pct(sid)

        if saebrs:
            screened_count += 1
            rl = saebrs.get("risk_level", "Low Risk")
            if rl == "High Risk":
                risk_distribution["high"] += 1
            elif rl == "Some Risk":
                risk_distribution["some"] += 1
            else:
                risk_distribution["low"] += 1

        if saebrs and plus:
            t = compute_mtss_tier(saebrs["risk_level"], plus["wellbeing_tier"], att_pct)
            tier_by_year[yr][f"tier{t}"] += 1
            class_breakdown[cls][f"tier{t}"] += 1
            tier_distribution[t] = tier_distribution.get(t, 0) + 1
            domain_totals["social"] += plus["social_domain"]
            domain_totals["academic"] += plus["academic_domain"]
            domain_totals["emotional"] += plus["emotional_domain"]
            domain_totals["belonging"] += plus["belonging_domain"]
            domain_totals["attendance"] += plus["attendance_domain"]
            domain_counts += 1

    domain_avgs = {k: round(v / domain_counts, 1) if domain_counts > 0 else 0 for k, v in domain_totals.items()}
    total = len(students)
    return {
        "tier_by_year": tier_by_year,
        "tier_distribution": tier_distribution,
        "risk_distribution": risk_distribution,
        "class_breakdown": class_breakdown,
        "domain_averages": domain_avgs,
        "total_students": total,
        "screened_students": screened_count,
        "screening_rate": round(screened_count / total * 100) if total > 0 else 0,
    }

@api_router.get("/analytics/cohort-comparison")
async def cohort_comparison(group_by: str = "year_level", user=Depends(get_current_user)):
    """Compare cohorts by year_level or class_name.
    Returns per-cohort: tier distribution, avg attendance, risk distribution, student count.
    """
    if group_by not in ("year_level", "class_name"):
        raise HTTPException(400, "group_by must be 'year_level' or 'class_name'")

    students = await db.students.find({"enrolment_status": "active"}, {"_id": 0}).to_list(500)
    cohorts: dict = {}

    for s in students:
        key = s.get(group_by) or "Unknown"
        if key not in cohorts:
            cohorts[key] = {"name": key, "total": 0, "screened": 0,
                            "tier1": 0, "tier2": 0, "tier3": 0,
                            "risk_low": 0, "risk_some": 0, "risk_high": 0,
                            "att_sum": 0.0}
        cohorts[key]["total"] += 1

        sid = s["student_id"]
        saebrs = await db.saebrs_results.find_one({"student_id": sid}, {"_id": 0}, sort=[("created_at", -1)])
        plus = await db.saebrs_plus_results.find_one({"student_id": sid}, {"_id": 0}, sort=[("created_at", -1)])
        att_pct = await get_student_attendance_pct(sid)
        cohorts[key]["att_sum"] += att_pct

        if saebrs:
            cohorts[key]["screened"] += 1
            rl = saebrs.get("risk_level", "Low Risk")
            if rl == "High Risk":
                cohorts[key]["risk_high"] += 1
            elif rl == "Some Risk":
                cohorts[key]["risk_some"] += 1
            else:
                cohorts[key]["risk_low"] += 1

        if saebrs and plus:
            t = compute_mtss_tier(saebrs["risk_level"], plus["wellbeing_tier"], att_pct)
        elif saebrs:
            t = 3 if saebrs["risk_level"] == "High Risk" else (2 if saebrs["risk_level"] == "Some Risk" else 1)
        else:
            t = None
        if t:
            cohorts[key][f"tier{t}"] += 1

    result = []
    for k, d in sorted(cohorts.items()):
        avg_att = round(d["att_sum"] / d["total"], 1) if d["total"] > 0 else 0
        result.append({
            "name": d["name"], "total": d["total"], "screened": d["screened"],
            "tier1": d["tier1"], "tier2": d["tier2"], "tier3": d["tier3"],
            "risk_low": d["risk_low"], "risk_some": d["risk_some"], "risk_high": d["risk_high"],
            "avg_attendance": avg_att,
        })
    return result

@api_router.get("/analytics/intervention-outcomes")
async def intervention_outcomes(user=Depends(get_current_user)):
    interventions = await db.interventions.find({}, {"_id": 0}).to_list(500)
    by_type = {}
    for i in interventions:
        t = i["intervention_type"]
        if t not in by_type:
            by_type[t] = {"total": 0, "completed": 0, "active": 0, "ratings": []}
        by_type[t]["total"] += 1
        if i["status"] == "completed":
            by_type[t]["completed"] += 1
            if i.get("outcome_rating"):
                by_type[t]["ratings"].append(i["outcome_rating"])
        elif i["status"] == "active":
            by_type[t]["active"] += 1
    return [{"type": t, "total": d["total"], "completed": d["completed"], "active": d["active"],
             "completion_rate": round(d["completed"] / d["total"] * 100) if d["total"] > 0 else 0,
             "avg_rating": round(sum(d["ratings"]) / len(d["ratings"]), 1) if d["ratings"] else None}
            for t, d in by_type.items()]

@api_router.get("/analytics/attendance-trends")
async def attendance_trends(user=Depends(get_current_user)):
    """Returns attendance trend data: monthly averages, chronic absentees, day-of-week patterns."""
    students = await db.students.find({"enrolment_status": "active"}, {"_id": 0}).to_list(500)
    settings_doc = await get_school_settings_doc()
    excluded_types = set(settings_doc.get("excluded_absence_types", []))
    school_days_list = await db.school_days.distinct("date")

    monthly: dict = {}
    chronic_absentees = []
    day_of_week: dict = {0: {"sessions": 0, "absent": 0}, 1: {"sessions": 0, "absent": 0},
                         2: {"sessions": 0, "absent": 0}, 3: {"sessions": 0, "absent": 0},
                         4: {"sessions": 0, "absent": 0}}

    for s in students:
        sid = s["student_id"]
        att_pct = await get_student_attendance_pct(sid)

        if att_pct < 90:
            chronic_absentees.append({
                "student": s, "attendance_pct": round(att_pct, 1),
                "tier": 3 if att_pct < 80 else 2
            })

        # Monthly data from exception records
        records = await db.attendance_records.find({"student_id": sid}, {"_id": 0}).to_list(3000)
        for r in records:
            month = r.get("date", "")[:7]
            if month:
                if month not in monthly:
                    monthly[month] = {"absent": 0, "total": 0}
                for sess in [r.get("am_status", ""), r.get("pm_status", "")]:
                    s_v = (sess or "").strip()
                    if s_v and s_v not in excluded_types:
                        monthly[month]["total"] += 1
                        if s_v not in PRESENT_STATUSES:
                            monthly[month]["absent"] += 1

            # Day of week pattern from exception records
            try:
                from datetime import date as date_obj
                d = date_obj.fromisoformat(r.get("date", ""))
                dow = d.weekday()  # 0=Mon, 4=Fri
                for sess in [r.get("am_status", ""), r.get("pm_status", "")]:
                    s_v = (sess or "").strip()
                    if s_v and s_v not in excluded_types:
                        day_of_week[dow]["total"] = day_of_week[dow].get("total", 0) + 1
                        if s_v not in PRESENT_STATUSES:
                            day_of_week[dow]["absent"] = day_of_week[dow].get("absent", 0) + 1
            except Exception:
                pass

    # For school_days-based data, supplement monthly with "present" sessions
    if school_days_list:
        for day in school_days_list:
            month = day[:7]
            if month not in monthly:
                monthly[month] = {"absent": 0, "total": 0}
            # Each school day = 2 sessions × students total (approx)
            # This gives a baseline for % calculation

    monthly_trend = [
        {"month": k, "absence_rate": round(v["absent"] / v["total"] * 100, 1) if v.get("total", 0) > 0 else 0}
        for k, v in sorted(monthly.items())
    ]

    DOW_NAMES = {0: "Monday", 1: "Tuesday", 2: "Wednesday", 3: "Thursday", 4: "Friday"}
    dow_trend = [
        {"day": DOW_NAMES[k], "absence_rate": round(v.get("absent", 0) / v.get("total", 1) * 100, 1)}
        for k, v in sorted(day_of_week.items()) if v.get("total", 0) > 0
    ]

    chronic_absentees.sort(key=lambda x: x["attendance_pct"])

    return {
        "monthly_trend": monthly_trend,
        "day_of_week": dow_trend,
        "chronic_absentees": chronic_absentees[:20],
        "total_school_days": len(school_days_list),
    }

# ==============================
# MEETING PREP
# ==============================

@api_router.get("/meeting-prep")
async def meeting_prep(user=Depends(get_current_user)):
    students = await db.students.find({"enrolment_status": "active"}, {"_id": 0}).to_list(500)
    result = []
    tier_changes = []

    for s in students:
        sid = s["student_id"]
        all_saebrs = await db.saebrs_results.find({"student_id": sid}, {"_id": 0}).sort("created_at", 1).to_list(10)
        saebrs = all_saebrs[-1] if all_saebrs else None
        prev_saebrs = all_saebrs[-2] if len(all_saebrs) >= 2 else None
        plus = await db.saebrs_plus_results.find_one({"student_id": sid}, {"_id": 0}, sort=[("created_at", -1)])
        att_pct = await get_student_attendance_pct(sid)

        if saebrs and plus:
            tier = compute_mtss_tier(saebrs["risk_level"], plus["wellbeing_tier"], att_pct)
        elif saebrs:
            tier = 3 if saebrs["risk_level"] == "High Risk" else (2 if saebrs["risk_level"] == "Some Risk" else 1)
        else:
            continue

        # Detect tier change between last two screenings
        if prev_saebrs:
            prev_tier = 3 if prev_saebrs["risk_level"] == "High Risk" else (2 if prev_saebrs["risk_level"] == "Some Risk" else 1)
            if prev_tier != tier:
                tier_changes.append({
                    "student": s,
                    "previous_tier": prev_tier,
                    "current_tier": tier,
                    "direction": "improved" if tier < prev_tier else "declined",
                    "previous_screening": prev_saebrs.get("created_at", ""),
                    "current_screening": saebrs.get("created_at", ""),
                    "saebrs": saebrs,
                })

        if tier >= 2:
            interventions = await db.interventions.find({"student_id": sid, "status": "active"}, {"_id": 0}).to_list(5)
            result.append({
                "student": s, "mtss_tier": tier,
                "saebrs": saebrs, "saebrs_plus": plus,
                "active_interventions": interventions,
                "attendance_pct": round(att_pct, 1)
            })

    result.sort(key=lambda x: -x["mtss_tier"])
    return {"students": result, "tier_changes": tier_changes}

# ==============================
# SETTINGS & DATA MANAGEMENT
# ==============================

@api_router.get("/public-settings")
async def public_settings():
    """No auth required — returns only branding fields for login/onboarding."""
    s = await db.school_settings.find_one({}, {"_id": 0})
    base = {k: SETTINGS_DEFAULTS[k] for k in ("platform_name", "accent_color", "logo_base64", "welcome_message", "school_name")}
    if s:
        for k in base:
            if s.get(k) is not None:
                base[k] = s[k]
    return base

@api_router.get("/school-days")
async def get_school_days(user=Depends(get_current_user)):
    """Returns list of registered school days (uploaded attendance dates)."""
    days = await db.school_days.distinct("date")
    return {"school_days": sorted(days), "total": len(days)}

@api_router.get("/settings")
async def get_settings(user=Depends(get_current_user)):
    s = await db.school_settings.find_one({}, {"_id": 0})
    if s:
        return {**SETTINGS_DEFAULTS, **s}
    return SETTINGS_DEFAULTS

@api_router.put("/settings")
async def update_settings(data: dict, user=Depends(get_current_user)):
    """Accept full settings dict — never overwrite onboarding_complete."""
    data.pop("_id", None)
    data.pop("onboarding_complete", None)
    await db.school_settings.update_one({}, {"$set": data}, upsert=True)
    result = await db.school_settings.find_one({}, {"_id": 0})
    return {**SETTINGS_DEFAULTS, **(result or {})}

@api_router.delete("/settings/data")
async def wipe_data(user=Depends(get_current_user)):
    for col in ["students", "attendance", "attendance_records", "school_days", "screening_sessions",
                "saebrs_results", "saebrs_plus_results", "interventions", "case_notes", "alerts"]:
        await db[col].delete_many({})
    return {"message": "All data wiped"}

@api_router.post("/settings/seed")
async def seed_data_endpoint(user=Depends(get_current_user)):
    return await seed_database()

@api_router.get("/classes")
async def get_classes(user=Depends(get_current_user)):
    classes = await db.students.distinct("class_name")
    result = []
    for cls in sorted(classes):
        s = await db.students.find_one({"class_name": cls}, {"_id": 0})
        result.append({"class_name": cls, "teacher": s["teacher"] if s else ""})
    return result

# ==============================
# ATTENDANCE MODULE
# ==============================

DEFAULT_ABSENCE_TYPES = [
    "Present", "Medical/Illness", "Unexplained", "Family Holiday",
    "Parent Choice School Approved", "Healthcare Appoint, Offsite",
    "Late arrival, approved", "Late arrival at School",
    "Early departure, approved", "Early departure from School",
    "School refusal or school can't", "Social services or justice",
    "Suspension - External", "Suspension - Internal",
]

@api_router.post("/attendance/upload")
async def upload_attendance(file: UploadFile = File(...), user=Depends(get_current_user)):
    """Exception-based upload: file contains only absences/exceptions.
    Students not in file for a given date are assumed fully present.
    'Present' status in the file = half-day attendance.
    """
    if user.get("role") not in ["admin", "leadership"]:
        raise HTTPException(403, "Admin or leadership access required")
    content = await file.read()
    fname = (file.filename or "").lower()
    records = []

    if fname.endswith('.xlsx') or fname.endswith('.xls'):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(cell.value or '').strip().upper() for cell in ws[1]]
        def col(names):
            for n in names:
                for i, h in enumerate(headers):
                    if h == n.upper() or h.startswith(n.upper()):
                        return i
            return None
        id_c = col(['SUSSIID', 'SUSSI ID', 'ID', 'STUDENT ID', 'STUDENTID', 'STUDENT CODE'])
        dt_c = col(['DATE'])
        am_c = col(['AM'])
        pm_c = col(['PM'])
        if id_c is None or dt_c is None:
            raise HTTPException(400, "Could not find ID or Date column")
        for row in ws.iter_rows(min_row=2, values_only=True):
            ext_id = str(row[id_c] or '').strip() if id_c is not None else ''
            dv = row[dt_c] if dt_c is not None else ''
            am_v = str(row[am_c] or '').strip() if am_c is not None else ''
            pm_v = str(row[pm_c] or '').strip() if pm_c is not None else ''
            if not ext_id or not dv:
                continue
            if hasattr(dv, 'strftime'):
                date_str = dv.strftime('%Y-%m-%d')
            else:
                ds = str(dv).strip()
                for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                    try:
                        date_str = datetime.strptime(ds, fmt).strftime('%Y-%m-%d'); break
                    except: pass
                else:
                    continue
            records.append({'external_id': ext_id.upper(), 'date': date_str, 'am_status': am_v, 'pm_status': pm_v})
    elif fname.endswith('.csv'):
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            ext_id = (row.get('SussiId') or row.get('SUSSIID') or row.get('ID') or row.get('Student ID') or row.get('student_id') or '').strip().upper()
            date_str = (row.get('Date') or row.get('date') or '').strip()
            am_v = (row.get('AM') or row.get('am') or '').strip()
            pm_v = (row.get('PM') or row.get('pm') or '').strip()
            if not ext_id or not date_str:
                continue
            for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
                try:
                    date_str = datetime.strptime(date_str, fmt).strftime('%Y-%m-%d'); break
                except: pass
            records.append({'external_id': ext_id, 'date': date_str, 'am_status': am_v, 'pm_status': pm_v})
    else:
        raise HTTPException(400, "Unsupported file format. Use XLSX or CSV.")

    if not records:
        raise HTTPException(400, "No valid records found in file")

    # Register school days (unique dates in this upload)
    unique_dates = list({r['date'] for r in records})
    for d in unique_dates:
        await db.school_days.update_one({"date": d}, {"$set": {"date": d}}, upsert=True)

    # Match students by sussi_id / external_id
    students_list = await db.students.find({}, {"_id": 0, "student_id": 1, "external_id": 1, "sussi_id": 1}).to_list(1000)
    ext_to_student = {}
    for s in students_list:
        if s.get("sussi_id"):
            ext_to_student[s["sussi_id"].upper()] = s["student_id"]
        if s.get("external_id"):
            ext_to_student[s["external_id"].upper()] = s["student_id"]

    by_ext = defaultdict(list)
    for r in records:
        by_ext[r['external_id']].append(r)

    matched, unmatched = [], []
    stored_count = 0
    for ext_id, recs in by_ext.items():
        student_id = ext_to_student.get(ext_id)
        if student_id:
            # Delete old records for matching dates only (allow additive uploads)
            for r in recs:
                await db.attendance_records.delete_one({"student_id": student_id, "date": r['date']})
            docs = [{"student_id": student_id, "external_id": ext_id, "date": r['date'],
                     "am_status": r['am_status'], "pm_status": r['pm_status']} for r in recs]
            if docs:
                await db.attendance_records.insert_many(docs)
                stored_count += len(docs)
            matched.append(ext_id)
        else:
            unmatched.append(ext_id)

    # Discover new absence types from this upload
    found_types = set()
    for r in records:
        for v in [r['am_status'], r['pm_status']]:
            if v:
                found_types.add(v)
    school_s = await db.school_settings.find_one({}, {"_id": 0})
    existing_types = set(school_s.get("absence_types") or DEFAULT_ABSENCE_TYPES)
    new_types = list(existing_types | found_types)
    await db.school_settings.update_one({}, {"$set": {"absence_types": sorted(new_types)}}, upsert=True)

    # Auto-generate attendance alerts for matched students
    alerts_created = 0
    student_map = {s["student_id"]: s for s in await db.students.find(
        {"student_id": {"$in": list(ext_to_student.values())}}, {"_id": 0}).to_list(500)}

    for ext_id in matched:
        sid = ext_to_student.get(ext_id)
        if not sid:
            continue
        s = student_map.get(sid, {})
        pref = s.get("preferred_name")
        display = f"{s.get('first_name', '')}{ (' (' + pref + ')') if pref else ''} {s.get('last_name', '')}".strip()
        att_pct = await get_student_attendance_pct(sid)

        # Determine alert severity
        if att_pct < 80:
            severity, alert_type = "high", "low_attendance_80"
            msg = f"{display} critically low attendance ({att_pct:.0f}%)"
        elif att_pct < 90:
            severity, alert_type = "medium", "low_attendance_90"
            msg = f"{display} attendance below 90% ({att_pct:.0f}%)"
        else:
            # Attendance OK — resolve any existing pending attendance alert
            await db.alerts.update_many(
                {"student_id": sid, "alert_type": {"$in": ["low_attendance_80", "low_attendance_90"]}, "status": "pending"},
                {"$set": {"resolved": True, "status": "resolved"}}
            )
            continue

        # Upsert: replace any existing pending attendance alert for this student
        await db.alerts.update_one(
            {"student_id": sid, "alert_type": {"$in": ["low_attendance_80", "low_attendance_90"]}, "status": "pending"},
            {"$set": {
                "alert_id": f"alrt_{uuid.uuid4().hex[:8]}",
                "student_id": sid, "student_name": display,
                "class_name": s.get("class_name", ""),
                "type": "early_warning", "alert_type": alert_type,
                "severity": severity, "message": msg,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "is_read": False, "resolved": False, "status": "pending"
            }},
            upsert=True
        )
        alerts_created += 1

    return {
        "processed": len(records),
        "school_days_registered": len(unique_dates),
        "matched_students": len(matched),
        "unmatched_students": len(unmatched),
        "stored_records": stored_count,
        "alerts_generated": alerts_created,
        "unmatched_ids": unmatched[:30],
    }

@api_router.get("/attendance/summary")
async def get_attendance_summary(user=Depends(get_current_user)):
    students_list = await db.students.find({"enrolment_status": "active"}, {"_id": 0}).to_list(500)
    school_days_list = await db.school_days.distinct("date")
    result = []
    for s in students_list:
        stats = await get_student_attendance_stats(s["student_id"])
        att_pct = stats["pct"]
        has_data = len(school_days_list) > 0 or stats["total_sessions"] > 0
        if not has_data:
            result.append({**s, "attendance_pct": None, "total_sessions": 0, "absent_sessions": 0, "has_data": False, "attendance_tier": None})
            continue
        tier = 1 if att_pct >= 95 else (2 if att_pct >= 90 else 3)
        result.append({**s, "attendance_pct": round(att_pct, 1), "total_sessions": stats["total_sessions"],
                       "absent_sessions": stats["absent_sessions"], "has_data": True, "attendance_tier": tier})
    return result

@api_router.get("/attendance/student/{student_id}")
async def get_student_attendance_detail(student_id: str, user=Depends(get_current_user)):
    records = await db.attendance_records.find({"student_id": student_id}, {"_id": 0}).sort("date", 1).to_list(3000)
    total_sessions, absent_sessions = 0, 0
    absence_types: dict = {}
    monthly_data: dict = {}
    for r in records:
        for sess_key, sess_v in [("am_status", r.get("am_status", "")), ("pm_status", r.get("pm_status", ""))]:
            v = (sess_v or "").strip()
            month = r.get("date", "")[:7]
            if v:
                total_sessions += 1
                if month not in monthly_data:
                    monthly_data[month] = {"sessions": 0, "absent": 0}
                monthly_data[month]["sessions"] += 1
                if v not in PRESENT_STATUSES:
                    absent_sessions += 1
                    absence_types[v] = absence_types.get(v, 0) + 1
                    monthly_data[month]["absent"] += 1
    att_pct = ((total_sessions - absent_sessions) / total_sessions * 100) if total_sessions > 0 else 100.0
    monthly_trend = [
        {"month": k, "attendance_pct": round(((v["sessions"] - v["absent"]) / v["sessions"] * 100) if v["sessions"] > 0 else 100, 1)}
        for k, v in sorted(monthly_data.items())
    ]
    return {
        "student_id": student_id, "attendance_pct": round(att_pct, 1),
        "total_sessions": total_sessions, "absent_sessions": absent_sessions,
        "absence_types": absence_types, "monthly_trend": monthly_trend,
        "records": records[:300],
    }

@api_router.get("/attendance/types")
async def get_absence_types(user=Depends(get_current_user)):
    s = await db.school_settings.find_one({}, {"_id": 0})
    return {"types": (s.get("absence_types") if s else None) or DEFAULT_ABSENCE_TYPES}

@api_router.put("/attendance/types")
async def update_absence_types(data: dict, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin access required")
    await db.school_settings.update_one({}, {"$set": {"absence_types": data.get("types", [])}}, upsert=True)
    return {"types": data.get("types", [])}

@api_router.put("/students/{student_id}/external-id")
async def set_student_external_id(student_id: str, data: dict, user=Depends(get_current_user)):
    if user.get("role") not in ["admin", "leadership"]:
        raise HTTPException(403, "Access denied")
    ext_id = data.get("external_id", "").strip().upper()
    await db.students.update_one({"student_id": student_id}, {"$set": {"external_id": ext_id}})
    return {"message": "External ID updated", "external_id": ext_id}

@api_router.put("/alerts/{alert_id}/approve")
async def approve_tier_change(alert_id: str, user=Depends(get_current_user)):
    if user.get("role") not in ["admin", "leadership", "wellbeing"]:
        raise HTTPException(403, "Access denied")
    await db.alerts.update_one({"alert_id": alert_id},
                                {"$set": {"pending_approval": False, "resolved": True, "is_read": True,
                                          "approved_by": user.get("user_id"), "approved_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "Tier change approved"}

@api_router.put("/alerts/{alert_id}/reject")
async def reject_tier_change(alert_id: str, user=Depends(get_current_user)):
    if user.get("role") not in ["admin", "leadership", "wellbeing"]:
        raise HTTPException(403, "Access denied")
    await db.alerts.update_one({"alert_id": alert_id},
                                {"$set": {"pending_approval": False, "rejected": True, "is_read": True}})
    return {"message": "Tier change rejected"}

# ==============================
# CSV REPORTS
# ==============================

@api_router.get("/reports/students-csv")
async def students_csv(user=Depends(get_current_user)):
    students = await db.students.find({}, {"_id": 0}).to_list(500)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Student ID", "First Name", "Last Name", "Year Level", "Class", "Teacher", "Gender", "Status"])
    for s in students:
        w.writerow([s["student_id"], s["first_name"], s["last_name"], s["year_level"], s["class_name"], s["teacher"], s.get("gender",""), s.get("enrolment_status","")])
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=students.csv"})

@api_router.get("/reports/tier-summary-csv")
async def tier_csv(user=Depends(get_current_user)):
    students = await db.students.find({}, {"_id": 0}).to_list(500)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Student ID", "First Name", "Last Name", "Class", "MTSS Tier", "SAEBRS Risk", "SAEBRS Total", "Wellbeing Tier", "Wellbeing Total", "Attendance %", "Active Interventions"])
    for s in students:
        sid = s["student_id"]
        saebrs = await db.saebrs_results.find_one({"student_id": sid}, {"_id": 0}, sort=[("created_at", -1)])
        plus = await db.saebrs_plus_results.find_one({"student_id": sid}, {"_id": 0}, sort=[("created_at", -1)])
        att_pct = await get_student_attendance_pct(sid)
        int_count = await db.interventions.count_documents({"student_id": sid, "status": "active"})
        if saebrs and plus:
            tier = compute_mtss_tier(saebrs["risk_level"], plus["wellbeing_tier"], att_pct)
        elif saebrs:
            tier = 3 if saebrs["risk_level"] == "High Risk" else (2 if saebrs["risk_level"] == "Some Risk" else 1)
        else:
            tier = "N/A"
        w.writerow([sid, s["first_name"], s["last_name"], s["class_name"],
                    f"Tier {tier}" if isinstance(tier, int) else tier,
                    saebrs["risk_level"] if saebrs else "Not Screened",
                    saebrs["total_score"] if saebrs else "",
                    f"Tier {plus['wellbeing_tier']}" if plus else "",
                    plus["wellbeing_total"] if plus else "",
                    f"{att_pct:.1f}", int_count])
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=tier_summary.csv"})

@api_router.get("/reports/screening-csv")
async def screening_csv(user=Depends(get_current_user)):
    results = await db.saebrs_results.find({}, {"_id": 0}).to_list(2000)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Student ID", "Screening ID", "Social", "Academic", "Emotional", "Total", "Risk Level", "Social Risk", "Academic Risk", "Emotional Risk", "Date"])
    for r in results:
        w.writerow([r["student_id"], r["screening_id"], r["social_score"], r["academic_score"], r["emotional_score"], r["total_score"], r["risk_level"], r["social_risk"], r["academic_risk"], r["emotional_risk"], r.get("created_at","")])
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=screening.csv"})

@api_router.get("/reports/interventions-csv")
async def interventions_csv(user=Depends(get_current_user)):
    items = await db.interventions.find({}, {"_id": 0}).to_list(500)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["ID", "Student ID", "Type", "Staff", "Start Date", "Review Date", "Status", "Goals", "Outcome Rating"])
    for i in items:
        w.writerow([i["intervention_id"], i["student_id"], i["intervention_type"], i["assigned_staff"], i["start_date"], i["review_date"], i["status"], i.get("goals",""), i.get("outcome_rating","")])
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=interventions.csv"})

@api_router.get("/settings/export-all")
async def export_all_data(user=Depends(get_current_user)):
    """Export ALL data as JSON backup — includes every collection."""
    backup = {}
    collections = [
        "students", "attendance_records", "school_days",
        "screening_sessions", "saebrs_results", "saebrs_plus_results",
        "interventions", "case_notes", "alerts", "school_settings",
    ]
    for col in collections:
        docs = await db[col].find({}, {"_id": 0}).to_list(50000)
        backup[col] = docs
    backup["exported_at"] = datetime.now(timezone.utc).isoformat()
    backup["version"] = "2.0"
    json_bytes = json.dumps(backup, default=str, indent=2).encode("utf-8")
    return StreamingResponse(
        iter([json_bytes]),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=welltrack_backup.json"}
    )

@api_router.post("/settings/restore")
async def restore_all_data(request: Request, user=Depends(get_current_user)):
    """Restore data from JSON backup — handles both v1.0 and v2.0 backups."""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    body = await request.json()
    # Support both old "attendance" key (v1) and new "attendance_records" key (v2)
    restorable = [
        "students", "attendance_records", "school_days",
        "screening_sessions", "saebrs_results", "saebrs_plus_results",
        "interventions", "case_notes", "alerts", "school_settings",
    ]
    restored = {}
    for col in restorable:
        source_key = col
        # v1 backups used "attendance" for what is now "attendance_records"
        if col == "attendance_records" and col not in body and "attendance" in body:
            source_key = "attendance"
        if source_key in body and isinstance(body[source_key], list):
            await db[col].delete_many({})
            if body[source_key]:
                await db[col].insert_many([{**doc} for doc in body[source_key]])
            restored[col] = len(body[source_key])
    return {"message": "Data restored successfully", "restored": restored}

# ==============================
# SEED FUNCTION
# ==============================


async def seed_database():
    for col in ["students", "attendance", "attendance_records", "school_days", "screening_sessions",
                "saebrs_results", "saebrs_plus_results", "interventions", "case_notes", "alerts"]:
        await db[col].delete_many({})

    await db.school_settings.update_one({}, {"$set": {
        "school_name": "Riverside Community School", "school_type": "both",
        "current_term": "Term 2", "current_year": 2025,
        # Seed excluded absence types so the feature is demonstrable
        "excluded_absence_types": ["Camp", "Excursion", "School Event"],
    }}, upsert=True)

    rng = random.Random(42)

    classes_data = [
        {"class": "3A", "teacher": "Ms Thompson", "year": "Year 3"},
        {"class": "5B", "teacher": "Mr Rodriguez", "year": "Year 5"},
        {"class": "7C", "teacher": "Ms Chen", "year": "Year 7"},
        {"class": "9A", "teacher": "Mr Williams", "year": "Year 9"},
    ]

    first_names = ["Emma", "Liam", "Olivia", "Noah", "Ava", "Ethan", "Sophia", "Mason",
                   "Isabella", "James", "Charlotte", "Alexander", "Amelia", "William", "Harper",
                   "Benjamin", "Evelyn", "Lucas", "Abigail", "Henry", "Emily", "Sebastian",
                   "Elizabeth", "Jack", "Sofia", "Owen", "Avery", "Theodore", "Ella", "Carter",
                   "Scarlett", "Jayden"]
    preferred_names = [None, None, None, "Ollie", None, None, None, None,
                       "Izzy", None, None, "Alex", None, "Will", None,
                       "Ben", None, None, "Abby", "Hank", None, "Seb",
                       "Liz", None, None, None, None, "Theo", None, None,
                       None, "Jay"]
    last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis",
                  "Wilson", "Taylor", "Anderson", "Thomas", "Jackson", "White", "Harris", "Martin",
                  "Thompson", "Moore", "Young", "Lee", "Walker", "Allen", "Hall", "Nguyen",
                  "Robinson", "King", "Wright", "Scott", "Torres", "Green"]

    students = []
    name_idx = 0
    for cls_data in classes_data:
        for i in range(8):
            fname = first_names[name_idx % len(first_names)]
            pname = preferred_names[name_idx % len(preferred_names)]
            lname = last_names[(name_idx + 7) % len(last_names)]
            sussi = f"YUS{name_idx+1:04d}"
            students.append({
                "student_id": f"stu_{uuid.uuid4().hex[:8]}",
                "first_name": fname, "preferred_name": pname, "last_name": lname,
                "year_level": cls_data["year"], "class_name": cls_data["class"],
                "teacher": cls_data["teacher"],
                "date_of_birth": "2014-01-01", "gender": "Male" if name_idx % 2 == 0 else "Female",
                "enrolment_status": "active",
                "sussi_id": sussi, "external_id": sussi,
            })
            name_idx += 1

    await db.students.insert_many(students)

    await db.screening_sessions.insert_many([
        {"screening_id": "scr_term1_2025", "screening_period": "Term 1", "year": 2025, "date": "2025-02-15", "teacher_id": "demo", "class_name": "all", "status": "completed"},
        {"screening_id": "scr_term2_2025", "screening_period": "Term 2", "year": 2025, "date": "2025-05-10", "teacher_id": "demo", "class_name": "all", "status": "completed"},
    ])

    # Build realistic school day calendars (weekdays only, skipping weekends)
    def build_school_days(start_date, num_days):
        days = []
        d = start_date
        while len(days) < num_days:
            if d.weekday() < 5:  # Mon–Fri only
                days.append(d.isoformat())
            d += timedelta(days=1)
        return days

    # Term 1: Feb–Mar, Term 2: Apr–May (creates monthly trend data across 4 months)
    term1_days = build_school_days(date_type(2025, 2, 3), 30)   # 30 days Term 1
    term2_days = build_school_days(date_type(2025, 4, 28), 25)  # 25 days Term 2
    all_school_days = term1_days + term2_days

    # SAEBRS item profiles
    low_s = [3,3,3,2,3,3]; low_a = [3,3,2,3,2,3]; low_e = [3,3,3,3,2,3,3]
    some_s = [2,2,2,2,1,2]; some_a = [2,2,1,2,1,2]; some_e = [2,2,2,1,2,2,2]
    high_s = [1,0,1,1,0,1]; high_a = [0,1,0,1,0,0]; high_e = [1,0,0,1,0,1,0]
    sr_low = [0,3,3,3,3,3,3]; sr_some = [2,2,2,2,2,2,2]; sr_high = [3,0,1,0,1,0,1]

    # Attendance percentages aligned to the 3-tier system:
    #   Tier 1 (≥95%): low risk students
    #   Tier 2 (90–95%): some risk students
    #   Tier 3 (<90%): high risk students
    att_map = {
        "low":  [97, 98, 96, 99],   # Tier 1: clearly above 95%
        "some": [91, 92, 93, 94],   # Tier 2: 90–95% band
        "high": [72, 68, 82, 75],   # Tier 3: well below 90%
    }

    # Students who change risk profile between Term 1 → Term 2 (creates tier changes)
    # idx: (term1_risk, term2_risk)
    risk_overrides = {
        3: ("some", "high"),   # Declined: some → high
        7: ("high", "some"),   # Improved: high → some
        11: ("low", "some"),   # Declined: low → some
        15: ("some", "low"),   # Improved: some → low
    }

    risk_cycle = ["low", "low", "low", "some", "some", "some", "high", "high"]

    def vary(items, delta=1):
        return [max(0, min(3, x + rng.randint(-delta, delta))) for x in items]

    # Absence types — realistic mix including some "neutral" types for excluded_absence_types demo
    regular_absences = [
        "Medical/Illness", "Medical/Illness", "Medical/Illness",   # most common
        "Unexplained", "Unexplained",
        "Family Holiday",
        "School refusal or school can't",
        "Parental choice",
        "Late arrival, approved",
        "Late arrival at School",
    ]
    neutral_absences = ["Camp", "Excursion", "School Event"]       # should be excluded

    case_note_templates = {
        3: [
            ("wellbeing",   "Wellbeing check-in. Student reported feeling overwhelmed and anxious about upcoming assessments. Discussed breathing techniques and time management strategies. Will follow up next week."),
            ("behaviour",   "Incident report: Student became distressed during group work and left the classroom. Spoke with student after — they shared they are struggling at home. Referred to school counsellor."),
            ("general",     "Follow-up meeting with parent/guardian. Family aware of school concerns. Parent noted similar behaviours at home. Collaborative support plan being drafted. Counsellor to lead."),
        ],
        2: [
            ("general",     "Regular check-in. Student appears to be managing but showing early signs of disengagement. Monitoring closely and checking in weekly. Encouraged to use check-in card."),
            ("wellbeing",   "Student self-referred to discuss friendship difficulties. Provided social skills strategies. Will observe in unstructured times and check in again next week."),
        ],
    }

    all_s1, all_s2, all_p1, all_p2 = [], [], [], []
    all_att_recs, all_int, all_notes, all_alerts = [], [], [], []
    school_days_set = set(all_school_days)

    for idx, student in enumerate(students):
        sid = student["student_id"]
        base_risk = risk_cycle[idx % len(risk_cycle)]

        # Apply risk overrides for tier change demo
        if idx in risk_overrides:
            risk_t1, risk_t2 = risk_overrides[idx]
        else:
            risk_t1 = risk_t2 = base_risk

        att_frac = rng.choice(att_map[base_risk]) / 100.0
        pref = student.get('preferred_name')
        display = f"{student['first_name']}{(' (' + pref + ')') if pref else ''} {student['last_name']}"

        def prof(risk):
            if risk == "low":   return low_s, low_a, low_e, sr_low
            if risk == "some":  return some_s, some_a, some_e, sr_some
            return high_s, high_a, high_e, sr_high

        s_t1, a_t1, e_t1, sr_t1 = prof(risk_t1)
        s_t2, a_t2, e_t2, sr_t2 = prof(risk_t2)

        def make_saebrs(s_items, a_items, e_items, screening_id, ts):
            s = sum(s_items); a = sum(a_items); e = sum(e_items); t = s + a + e
            r, sr, ar, er = compute_saebrs_risk(t, s, a, e)
            return {"result_id": str(uuid.uuid4()), "student_id": sid, "screening_id": screening_id,
                    "social_items": s_items, "academic_items": a_items, "emotional_items": e_items,
                    "social_score": s, "academic_score": a, "emotional_score": e, "total_score": t,
                    "risk_level": r, "social_risk": sr, "academic_risk": ar, "emotional_risk": er, "created_at": ts}

        def make_plus(saebrs_doc, sr_items, screening_id, att, ts):
            soc = saebrs_doc["social_score"]; aca = saebrs_doc["academic_score"]
            rev = 3 - sr_items[0]
            emo = rev + sr_items[1] + sr_items[2]
            bel = sr_items[3] + sr_items[4] + sr_items[5] + sr_items[6]
            att_s = compute_attendance_score(att * 100)
            att_d = att_s * 3
            total = soc + aca + emo + bel + att_d
            return {"result_id": str(uuid.uuid4()), "student_id": sid, "screening_id": screening_id,
                    "self_report_items": sr_items, "attendance_pct": att * 100,
                    "social_domain": soc, "academic_domain": aca, "emotional_domain": emo,
                    "belonging_domain": bel, "attendance_domain": att_d,
                    "wellbeing_total": total, "wellbeing_tier": compute_wellbeing_tier(total), "created_at": ts}

        s1 = make_saebrs(vary(s_t1), vary(a_t1), vary(e_t1), "scr_term1_2025", "2025-02-15T09:00:00")
        s2 = make_saebrs(vary(s_t2, 2), vary(a_t2, 2), vary(e_t2, 2), "scr_term2_2025", "2025-05-10T09:00:00")
        all_s1.append(s1); all_s2.append(s2)

        p1 = make_plus(s1, vary(sr_t1), "scr_term1_2025", att_frac, "2025-02-15T10:00:00")
        p2 = make_plus(s2, vary(sr_t2, 2), "scr_term2_2025", att_frac, "2025-05-10T10:00:00")
        all_p1.append(p1); all_p2.append(p2)

        # ── ATTENDANCE (exception-based, weekdays only, randomly distributed) ──
        total_days = len(all_school_days)
        absent_days_count = int(total_days * (1 - att_frac))

        # Randomly pick which days this student is absent
        absent_indices = set(rng.sample(range(total_days), min(absent_days_count, total_days)))

        # A few days: Camp/Excursion for excluded type demo (only for low-risk students)
        neutral_indices = set()
        if base_risk == "low" and total_days > 5:
            neutral_indices = set(rng.sample(range(total_days), rng.randint(1, 3)))

        for day_num, rec_date in enumerate(all_school_days):
            if day_num in neutral_indices:
                # Excluded absence type (Camp/Excursion)
                ntype = rng.choice(neutral_absences)
                all_att_recs.append({
                    "student_id": sid, "external_id": student.get("sussi_id", ""),
                    "date": rec_date, "am_status": ntype, "pm_status": ntype
                })
            elif day_num in absent_indices:
                abs_type = rng.choice(regular_absences)
                # ~20% of absences are half-day ("Present" for one session)
                if rng.random() < 0.20 and abs_type in ("Late arrival, approved", "Late arrival at School"):
                    all_att_recs.append({
                        "student_id": sid, "external_id": student.get("sussi_id", ""),
                        "date": rec_date, "am_status": abs_type, "pm_status": "Present"
                    })
                else:
                    all_att_recs.append({
                        "student_id": sid, "external_id": student.get("sussi_id", ""),
                        "date": rec_date, "am_status": abs_type, "pm_status": abs_type
                    })

        # ── TIERS ──
        final_tier = compute_mtss_tier(s2["risk_level"], p2["wellbeing_tier"], att_frac * 100)
        prev_tier  = compute_mtss_tier(s1["risk_level"], p1["wellbeing_tier"], att_frac * 100)

        # ── ALERTS ──
        if s2["risk_level"] == "High Risk":
            all_alerts.append({
                "alert_id": f"alrt_{uuid.uuid4().hex[:8]}", "student_id": sid,
                "student_name": display, "class_name": student["class_name"],
                "type": "early_warning", "alert_type": "high_risk_saebrs", "severity": "high",
                "message": f"{display} screened High Risk (SAEBRS: {s2['total_score']}/57)",
                "created_at": "2025-05-10T09:30:00", "is_read": False, "resolved": False, "status": "pending"
            })

        if att_frac * 100 < 80:
            all_alerts.append({
                "alert_id": f"alrt_{uuid.uuid4().hex[:8]}", "student_id": sid,
                "student_name": display, "class_name": student["class_name"],
                "type": "early_warning", "alert_type": "low_attendance_80", "severity": "high",
                "message": f"{display} critically low attendance ({att_frac*100:.0f}%)",
                "created_at": "2025-05-01T08:00:00", "is_read": False, "resolved": False, "status": "pending"
            })
        elif att_frac * 100 < 90:
            all_alerts.append({
                "alert_id": f"alrt_{uuid.uuid4().hex[:8]}", "student_id": sid,
                "student_name": display, "class_name": student["class_name"],
                "type": "early_warning", "alert_type": "low_attendance_90", "severity": "medium",
                "message": f"{display} attendance below 90% ({att_frac*100:.0f}%)",
                "created_at": "2025-05-01T08:00:00", "is_read": False, "resolved": False, "status": "pending"
            })

        if prev_tier != final_tier:
            all_alerts.append({
                "alert_id": f"alrt_{uuid.uuid4().hex[:8]}", "student_id": sid,
                "student_name": display, "class_name": student["class_name"],
                "type": "tier_change", "alert_type": "tier_change",
                "from_tier": prev_tier, "to_tier": final_tier,
                "severity": "high" if final_tier > prev_tier else "medium",
                "message": f"{display} moved from Tier {prev_tier} to Tier {final_tier}",
                "created_at": "2025-05-10T09:30:00", "is_read": False, "resolved": False, "status": "pending"
            })

        # ── INTERVENTIONS & CASE NOTES ──
        staff_options = ["Ms Parker (Wellbeing)", "Mr Lewis (Counsellor)", "Ms Ahmed (SENCO)", student["teacher"]]
        int_types_t3 = ["Counselling", "Behaviour Support", "Social Skills Groups", "Check-In/Check-Out"]
        int_types_t2 = ["Mentoring", "Academic Support", "Attendance Intervention", "Parent Consultation"]

        if final_tier == 3:
            # Active intervention from Term 2
            all_int.append({
                "intervention_id": f"int_{uuid.uuid4().hex[:8]}", "student_id": sid,
                "intervention_type": rng.choice(int_types_t3), "assigned_staff": rng.choice(staff_options[:3]),
                "start_date": "2025-04-28", "review_date": "2025-06-09", "status": "active",
                "goals": "Reduce risk indicators, build coping strategies and improve school connectedness.",
                "progress_notes": "Initial engagement established. Student is attending sessions. Monitoring weekly progress.",
                "frequency": "3x weekly", "outcome_rating": None, "created_at": "2025-04-28T09:00:00"
            })
            # Completed intervention from Term 1 (demonstrates analytics completion rates)
            all_int.append({
                "intervention_id": f"int_{uuid.uuid4().hex[:8]}", "student_id": sid,
                "intervention_type": rng.choice(int_types_t3), "assigned_staff": rng.choice(staff_options[:3]),
                "start_date": "2025-02-03", "review_date": "2025-03-28", "status": "completed",
                "goals": "Establish rapport and identify key risk factors.",
                "progress_notes": "Completed 8-week programme. Student showed moderate improvement in social engagement.",
                "frequency": "Weekly", "outcome_rating": rng.choice([3, 4, 4]), "created_at": "2025-02-03T09:00:00"
            })
            # Multiple case notes over time for Tier 3
            for tmpl_idx, (ntype, ntext) in enumerate(case_note_templates.get(3, [])):
                note_dates = ["2025-03-10", "2025-04-14", "2025-05-12"]
                all_notes.append({
                    "case_id": f"case_{uuid.uuid4().hex[:8]}", "student_id": sid,
                    "staff_member": rng.choice(["Ms Parker (Wellbeing)", student["teacher"], "Mr Lewis (Counsellor)"]),
                    "date": note_dates[tmpl_idx % len(note_dates)], "note_type": ntype, "notes": ntext,
                    "created_at": f"2025-0{3+tmpl_idx}-10T11:00:00"
                })

        elif final_tier == 2:
            # Active intervention
            all_int.append({
                "intervention_id": f"int_{uuid.uuid4().hex[:8]}", "student_id": sid,
                "intervention_type": rng.choice(int_types_t2), "assigned_staff": rng.choice(staff_options),
                "start_date": "2025-04-28", "review_date": "2025-06-09", "status": "active",
                "goals": "Monitor emerging risk factors and build protective factors through regular check-ins.",
                "progress_notes": "Weekly check-ins established. Student is receptive to support.",
                "frequency": "Weekly", "outcome_rating": None, "created_at": "2025-04-28T09:00:00"
            })
            # Completed intervention from Term 1
            if rng.random() < 0.5:
                all_int.append({
                    "intervention_id": f"int_{uuid.uuid4().hex[:8]}", "student_id": sid,
                    "intervention_type": rng.choice(int_types_t2), "assigned_staff": rng.choice(staff_options),
                    "start_date": "2025-02-17", "review_date": "2025-03-28", "status": "completed",
                    "goals": "Early monitoring and prevention of further risk escalation.",
                    "progress_notes": "6-week monitoring programme complete. No escalation observed.",
                    "frequency": "Fortnightly", "outcome_rating": rng.choice([3, 4, 5]), "created_at": "2025-02-17T09:00:00"
                })
            # Case notes for Tier 2
            for tmpl_idx, (ntype, ntext) in enumerate(case_note_templates.get(2, [])):
                note_dates = ["2025-04-07", "2025-05-08"]
                all_notes.append({
                    "case_id": f"case_{uuid.uuid4().hex[:8]}", "student_id": sid,
                    "staff_member": rng.choice([student["teacher"], "Ms Parker (Wellbeing)"]),
                    "date": note_dates[tmpl_idx % len(note_dates)], "note_type": ntype, "notes": ntext,
                    "created_at": f"2025-0{4+tmpl_idx}-07T09:30:00"
                })

    # Insert school_days (weekdays only, both terms)
    await db.school_days.insert_many([{"date": d} for d in sorted(school_days_set)])

    for col_data, col_name in [
        (all_s1 + all_s2, "saebrs_results"),
        (all_p1 + all_p2, "saebrs_plus_results"),
        (all_att_recs, "attendance_records"),
        (all_int, "interventions"),
        (all_notes, "case_notes"),
        (all_alerts, "alerts"),
    ]:
        if col_data:
            await db[col_name].insert_many(col_data)

    return {
        "message": "Demo data seeded",
        "students": len(students),
        "school_days": len(school_days_set),
        "attendance_records": len(all_att_recs),
        "interventions": len(all_int),
        "case_notes": len(all_notes),
        "alerts": len(all_alerts),
    }
# ==============================
# APP SETUP
# ==============================

app.include_router(api_router)

app.add_middleware(CORSMiddleware, allow_credentials=True, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
app.add_middleware(SessionMiddleware, secret_key=os.environ['SESSION_SECRET'])

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

@app.on_event("startup")
async def startup():
    logger.info("WellTrack starting up — onboarding flow active, no auto-seeding.")

@app.on_event("shutdown")
async def shutdown():
    client.close()
